from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.logage2026.config import CHARTS_DIR, Q11_WORKBOOK_OUTPUT


THIN_BORDER = Border(
    left=Side(style="thin", color="D9DEE5"),
    right=Side(style="thin", color="D9DEE5"),
    top=Side(style="thin", color="D9DEE5"),
    bottom=Side(style="thin", color="D9DEE5"),
)
TITLE_FILL = PatternFill("solid", fgColor="0F4C81")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="EAF2F8")
CLASS_FILLS = {
    "A": PatternFill("solid", fgColor="DDEBF7"),
    "B": PatternFill("solid", fgColor="FFF2CC"),
    "C": PatternFill("solid", fgColor="F4CCCC"),
    "X": PatternFill("solid", fgColor="D9EAD3"),
    "Y": PatternFill("solid", fgColor="FCE5CD"),
    "Z": PatternFill("solid", fgColor="F4CCCC"),
}
TITLE_FONT = Font(bold=True, color="FFFFFF", size=14)
SECTION_FONT = Font(bold=True, color="1F1F1F", size=11)
HEADER_FONT = Font(bold=True, color="FFFFFF")
BOLD_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write_dataframe_sheet(ws, title_text: str, df: pd.DataFrame) -> None:
    ws.column_dimensions[get_column_letter(1)].width = 3
    for col in range(2, len(df.columns) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 18
    _apply_title(ws, "B1", len(df.columns) + 1, title_text)
    
    for col, value in enumerate(df.columns, start=2):
        ws.cell(2, col, str(value))
    _style_range(ws, 2, 2, len(df.columns) + 1, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    
    for row_idx, row in enumerate(df.itertuples(index=False), start=3):
        for col_idx, value in enumerate(row, start=2):
            val_to_write = None if pd.isna(value) else value
            cell = ws.cell(row_idx, col_idx, val_to_write)
            cell.border = THIN_BORDER
            cell.alignment = LEFT if isinstance(val_to_write, str) else CENTER
            if isinstance(val_to_write, float):
                cell.number_format = "0.00"
    
    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"B2:{get_column_letter(len(df.columns) + 1)}{len(df) + 2}"


def write_summary_workbook(
    abc_xyz: pd.DataFrame,
    abc_xyz_matrix: pd.DataFrame,
    monthly_demand: pd.DataFrame,
    q11_shipments: pd.DataFrame,
    warehouse_region_summary: pd.DataFrame,
    q12_top_demand_provinces_summary: pd.DataFrame,
    warehouse_imbalance_summary: pd.DataFrame,
    q13_segment_profile_summary: pd.DataFrame,
    q13_segment_packaging_summary: pd.DataFrame,
    q13_segment_geographic_spread_summary: pd.DataFrame,
    safety_stock_class_a: pd.DataFrame,
    lead_time_sensitivity: pd.DataFrame,
    inventory_pooling_summary: pd.DataFrame,
    hcm_district_summary: pd.DataFrame,
    network_model_evaluation: pd.DataFrame,
    shipments: pd.DataFrame,
    slotting_plan: pd.DataFrame,
    sku_pick_profile: pd.DataFrame,
    travel_metrics: dict,
    lead_time_table: pd.DataFrame,
    output_path: Path,
) -> Path:
    workbook = Workbook()

    # ── Part 1 — Q1.1 ────────────────────────────────────────────────────────
    q11_ws = workbook.active
    q11_ws.title = "Q1.1 ABC-XYZ"
    _write_q11_abc_xyz_sheet(q11_ws, abc_xyz, abc_xyz_matrix, q11_shipments)
    _write_full_sku_ranking(workbook.create_sheet("Q1.1 Full SKU Ranking"), abc_xyz)

    # ── Part 1 — Q1.2 ────────────────────────────────────────────────────────
    _write_q12_heatmap_summary_sheet(
        workbook.create_sheet("Q1.2 Heatmap Summary"),
        warehouse_region_summary,
        q12_top_demand_provinces_summary,
        warehouse_imbalance_summary,
    )

    # ── Part 1 — Q1.3 ────────────────────────────────────────────────────────
    _write_q13_order_profile_summary_sheet(
        workbook.create_sheet("Q1.3 Order Profile Summary"),
        q13_segment_profile_summary,
        q13_segment_packaging_summary,
    )

    # ── Part 2 — Q2.1 ────────────────────────────────────────────────────────
    _write_q21_dark_store_sla_sheet(workbook.create_sheet("Q2.1 Dark Store SLA & Districts"), network_model_evaluation, shipments)

    # ── Part 2 — Q2.2 ────────────────────────────────────────────────────────
    _write_q22_lead_time_sheet(workbook.create_sheet("Q2.2 Lead Time"), lead_time_table)
    _write_q22_inventory_pooling_sheet(workbook.create_sheet("Q2.2 Safety Stock & Pooling"), inventory_pooling_summary, safety_stock_class_a)

    # ── Part 3 — Q3.1 ────────────────────────────────────────────────────────
    _write_q31_slotting_summary_sheet(workbook.create_sheet("Q3.1 Slotting Summary"), travel_metrics)
    _write_q31_slot_assignment_sheet(workbook.create_sheet("Q3.1 Slot Assignment"), slotting_plan)
    _write_q31_u_shape_sheet(workbook.create_sheet("Q3.1 Heatmap (U-shape)"))

    # ── Part 3 — Q3.2 ────────────────────────────────────────────────────────
    _write_q32_pick_pack_sheet(workbook.create_sheet("Q3.2 Pick & Pack"))

    workbook.save(output_path)
    return output_path


def _apply_title(ws, start_cell: str, end_col: int, text: str) -> None:
    row = ws[start_cell].row
    start_col = ws[start_cell].column
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws[start_cell]
    cell.value = text
    cell.fill = TITLE_FILL
    cell.font = TITLE_FONT
    cell.alignment = LEFT


def _style_range(ws, row: int, start_col: int, end_col: int, *, fill=None, font=None, alignment=None) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.border = THIN_BORDER
        if fill is not None:
            cell.fill = fill
        if font is not None:
            cell.font = copy(font)
        if alignment is not None:
            cell.alignment = alignment


def _format_share(value: float) -> str:
    return f"{value:.1%}"


def _safe_float(value: object) -> float:
    return 0.0 if pd.isna(value) else float(value)


def _period_label(start: pd.Timestamp, end: pd.Timestamp) -> str:
    if start.year == end.year:
        return f"{start.strftime('%b')}–{end.strftime('%b %Y')}"
    return f"{start.strftime('%b %Y')}–{end.strftime('%b %Y')}"


def _month_columns(monthly_demand: pd.DataFrame) -> list[str]:
    month_like = []
    for column in monthly_demand.columns:
        if column in {"sku_code", "product_name", "category", "abc_quantity", "xyz", "abc_xyz", "Total", "Mean/Mo", "Std Dev", "CV", "quantity", "quantity_share", "order_frequency", "frequency_share", "cbm_total"}:
            continue
        try:
            pd.Period(column, freq="M")
            month_like.append(column)
        except Exception:
            continue
    return month_like


def _operational_category(abc_xyz_code: str) -> str:
    mapping = {
        "AX": "Fast & Predictable",
        "AY": "Fast & Predictable",
        "AZ": "Fast & Volatile",
        "BX": "Medium & Stable",
        "BY": "Medium & Seasonal",
        "BZ": "Medium & Volatile",
        "CX": "Slow & Stable",
        "CY": "Slow & Seasonal",
        "CZ": "Slow & Erratic",
    }
    return mapping.get(abc_xyz_code, "Unclassified")


def _slotting_zone(abc_class: str) -> str:
    if abc_class == "A":
        return "Pick-Face Zone (Ground Level)"
    if abc_class == "B":
        return "Forward Reserve Zone"
    return "Reserve / Bulk Zone"


def _demand_pattern(xyz_class: str) -> str:
    return {"X": "Stable", "Y": "Seasonal", "Z": "Erratic"}.get(xyz_class, "Unknown")


def _stock_strategy(xyz_class: str) -> str:
    return {
        "X": "Fixed ROP + EOQ",
        "Y": "Dynamic MRP + safety stock",
        "Z": "Buffer stock + safety stock",
    }.get(xyz_class, "Manual review")


def _reorder_logic(xyz_class: str) -> str:
    return {"X": "Min-Max fixed", "Y": "Weekly forecast", "Z": "Safety stock + review"}.get(xyz_class, "Manual review")


def _dashboard_abc_summary(abc_xyz: pd.DataFrame) -> pd.DataFrame:
    summary = (
        abc_xyz.groupby("abc_quantity", dropna=False)
        .agg(sku_count=("sku_code", "nunique"), total_qty=("quantity", "sum"), total_cbm=("cbm_total", "sum"))
        .reset_index()
    )
    summary["sku_share"] = summary["sku_count"] / abc_xyz["sku_code"].nunique()
    summary["qty_share"] = summary["total_qty"] / abc_xyz["quantity"].sum()
    return summary.set_index("abc_quantity").reindex(["A", "B", "C"]).reset_index()


def _dashboard_xyz_frequency_summary(abc_xyz: pd.DataFrame) -> pd.DataFrame:
    summary = (
        abc_xyz.groupby("abc_frequency", dropna=False)
        .agg(sku_count=("sku_code", "nunique"), total_freq=("order_frequency", "sum"))
        .reset_index()
    )
    summary["sku_share"] = summary["sku_count"] / abc_xyz["sku_code"].nunique()
    return summary.set_index("abc_frequency").reindex(["A", "B", "C"]).reset_index()

def _dashboard_xyz_volatility_summary(abc_xyz: pd.DataFrame) -> pd.DataFrame:
    summary = (
        abc_xyz.groupby("xyz_volatility", dropna=False)
        .agg(sku_count=("sku_code", "nunique"), avg_cv=("demand_cv", "mean"))
        .reset_index()
    )
    summary["sku_share"] = summary["sku_count"] / abc_xyz["sku_code"].nunique()
    return summary.set_index("xyz_volatility").reindex(["X", "Y", "Z"]).reset_index()


def _write_q11_abc_xyz_sheet(
    ws,
    abc_xyz: pd.DataFrame,
    abc_xyz_matrix: pd.DataFrame,
    q11_shipments: pd.DataFrame,
) -> None:
    """Write the Q1.1 ABC-XYZ sheet matching the reference 6-section layout.

    Sections: A = ABC Quantity, B = ABC Frequency, C = XYZ Volatility,
              D = ABC-XYZ (Qty × Vol) matrix, E = ABC Qty × Freq matrix,
              F = Fast-Moving group (AA class).
    Charts are embedded below the data.
    """
    period_start = q11_shipments["created_date"].min()
    period_end = q11_shipments["created_date"].max()
    month_count = q11_shipments["created_date"].dt.to_period("M").nunique()
    for idx, width in enumerate([3, 18, 14, 12, 16, 14, 12, 24, 40], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    _apply_title(ws, "B1", 9, "LOGage 2026 — QUESTION 1.1 | ABC-XYZ ANALYSIS")
    period_text = (
        f"Data: My Phuoc & Vinh Loc Warehouses  |  Period: {_period_label(period_start, period_end)}  |  "
        f"{len(q11_shipments):,} transaction lines  |  {abc_xyz['sku_code'].nunique():,} unique SKUs  |  "
        f"Formula: ABC Qty/Freq (cum. 70/20/10%), XYZ Variability (CV ≤0.5/≤1.0/>1.0)"
    )
    ws.merge_cells("B2:I2")
    ws["B2"] = period_text
    ws["B2"].alignment = LEFT
    ws["B2"].font = Font(italic=True, color="5B5B5B")

    r = 4
    # ── A. ABC CLASSIFICATION — QUANTITY ──────────────────────────────────
    abc_summary = _dashboard_abc_summary(abc_xyz)
    ws.merge_cells(f"B{r}:I{r}")
    ws[f"B{r}"] = "A.  ABC CLASSIFICATION — OUTBOUND QUANTITY (threshold: A=70%, B=90%, C=100%)"
    _style_range(ws, r, 2, 9, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    r += 1
    headers = ["Class", "SKU Count", "% of SKUs", "Total Qty (pcs)", "% of Total Qty", "Total CBM", "Threshold", "Action"]
    for col, value in enumerate(headers, start=2):
        ws.cell(r, col, value)
    _style_range(ws, r, 2, 9, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    r += 1
    thresholds = {"A": "Cum. top 70% volume", "B": "70–90% cum. volume", "C": "Bottom 10% volume"}
    actions = {
        "A": "Pick-face zone | Daily cycle count | Min lead time",
        "B": "Intermediate zone | Weekly review",
        "C": "Reserve/bulk zone | Monthly review",
    }
    for row in abc_summary.itertuples(index=False):
        values = [row.abc_quantity, int(row.sku_count), _format_share(row.sku_share), round(_safe_float(row.total_qty)), _format_share(row.qty_share), round(_safe_float(row.total_cbm)), thresholds[row.abc_quantity], actions[row.abc_quantity]]
        for col, value in enumerate(values, start=2):
            ws.cell(r, col, value)
        _style_range(ws, r, 2, 9, fill=CLASS_FILLS[row.abc_quantity], alignment=LEFT)
        r += 1

    r += 1
    # ── B. ABC CLASSIFICATION — ORDER FREQUENCY ───────────────────────────
    xyz_freq_summary = _dashboard_xyz_frequency_summary(abc_xyz)
    ws.merge_cells(f"B{r}:I{r}")
    ws[f"B{r}"] = "B.  ABC CLASSIFICATION — ORDER FREQUENCY (threshold: A=70%, B=90%, C=100%)"
    _style_range(ws, r, 2, 9, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    r += 1
    freq_headers = ["Class", "SKU Count", "% of SKUs", "Total Freq", "Freq/SKU", "Demand Pattern", "Replenishment Logic", "Safety Stock"]
    for col, value in enumerate(freq_headers, start=2):
        ws.cell(r, col, value)
    _style_range(ws, r, 2, 9, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    r += 1
    demand = {"A": "High Frequency", "B": "Medium Frequency", "C": "Low Frequency"}
    repl = {"A": "Fixed ROP", "B": "Dynamic MRP/forecast", "C": "Demand-driven + buffer"}
    safety = {"A": "Low — 1–2 weeks cover", "B": "Medium — 2–4 weeks cover", "C": "High — 4–8 weeks cover"}
    meaning = {"A": "Pick nhiều nhất → slot gần đóng gói", "B": "Tần suất vừa → zone trung gian", "C": "Hiếm pick → zone dự trữ"}
    for row in xyz_freq_summary.itertuples(index=False):
        cls = getattr(row, "abc_frequency")
        freq_per_sku = _safe_float(row.total_freq) / int(row.sku_count) if int(row.sku_count) else 0.0
        values = [cls, int(row.sku_count), _format_share(row.sku_share), round(_safe_float(row.total_freq)), round(freq_per_sku, 1), demand[cls], repl[cls], safety[cls]]
        for col, value in enumerate(values, start=2):
            ws.cell(r, col, value)
        _style_range(ws, r, 2, 9, fill=CLASS_FILLS[cls], alignment=LEFT)
        r += 1

    r += 1
    # ── C. XYZ VOLATILITY CLASSIFICATION ─────────────────────────────────
    xyz_vol_summary = _dashboard_xyz_volatility_summary(abc_xyz)
    ws.merge_cells(f"B{r}:I{r}")
    ws[f"B{r}"] = f"C.  XYZ CLASSIFICATION — DEMAND VARIABILITY (CV over {month_count} months; X≤0.50, Y≤1.00, Z>1.00)"
    _style_range(ws, r, 2, 9, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    r += 1
    xyz_headers = ["Class", "CV Range", "SKU Count", "% of SKUs", "Demand Pattern", "Safety Stock", "Replenishment Logic", "Avg CV"]
    for col, value in enumerate(xyz_headers, start=2):
        ws.cell(r, col, value)
    _style_range(ws, r, 2, 9, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    r += 1
    cv_ranges = {"X": "CV ≤ 0.50", "Y": "0.50 < CV ≤ 1.00", "Z": "CV > 1.00"}
    demand_vol = {"X": "Stable / consistent", "Y": "Seasonal / trend", "Z": "Erratic / unpredictable"}
    repl_vol = {"X": "Fixed ROP + EOQ", "Y": "Dynamic MRP / forecast", "Z": "Buffer stock + safety stock"}
    safety_vol = {"X": "Thấp — 1–2 tuần", "Y": "TB — 2–4 tuần", "Z": "Cao — 4–8 tuần"}
    for row in xyz_vol_summary.itertuples(index=False):
        cls = getattr(row, "xyz_volatility")
        values = [cls, cv_ranges[cls], int(row.sku_count), _format_share(row.sku_share), demand_vol[cls], safety_vol[cls], repl_vol[cls], round(_safe_float(row.avg_cv), 2)]
        for col, value in enumerate(values, start=2):
            ws.cell(r, col, value)
        _style_range(ws, r, 2, 9, fill=CLASS_FILLS[cls], alignment=LEFT)
        r += 1

    r += 1
    # ── D. ABC-XYZ (Qty × Volatility) MATRIX ─────────────────────────────
    ws.merge_cells(f"B{r}:I{r}")
    ws[f"B{r}"] = "D.  ABC-XYZ MATRIX (Qty × Volatility)"
    _style_range(ws, r, 2, 9, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    r += 1
    matrix_vol = (
        abc_xyz.groupby(["abc_quantity", "xyz_volatility"]).size().unstack(fill_value=0)
        .reindex(index=["A", "B", "C"], columns=["X", "Y", "Z"])
        .fillna(0).astype(int)
    )
    headers = ["", "X (Stable)", "Y (Seasonal)", "Z (Erratic)", "Row Total", "% Volume", "% Freq", "Key Implication"]
    for col, value in enumerate(headers, start=2):
        ws.cell(r, col, value)
    _style_range(ws, r, 2, 9, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    r += 1
    volume_share = abc_xyz.groupby("abc_quantity")["quantity"].sum() / abc_xyz["quantity"].sum()
    freq_share = abc_xyz.groupby("abc_quantity")["order_frequency"].sum() / abc_xyz["order_frequency"].sum()
    implications = {
        "A": "HIGH PRESSURE — Daily ops, premium slots, SLA-critical",
        "B": "MODERATE — Standard fulfillment, replenishment planning",
        "C": "LOW PRIORITY — Bulk storage, periodic review",
    }
    for abc_class in ["A", "B", "C"]:
        values = [abc_class, int(matrix_vol.loc[abc_class, "X"]), int(matrix_vol.loc[abc_class, "Y"]), int(matrix_vol.loc[abc_class, "Z"]), int(matrix_vol.loc[abc_class].sum()), _format_share(_safe_float(volume_share.get(abc_class))), _format_share(_safe_float(freq_share.get(abc_class))), implications[abc_class]]
        for col, value in enumerate(values, start=2):
            ws.cell(r, col, value)
        _style_range(ws, r, 2, 9, fill=CLASS_FILLS[abc_class], alignment=LEFT)
        r += 1
    total_values = ["TOTAL", int(matrix_vol["X"].sum()), int(matrix_vol["Y"].sum()), int(matrix_vol["Z"].sum()), int(matrix_vol.values.sum()), "", "", ""]
    for col, value in enumerate(total_values, start=2):
        ws.cell(r, col, value)
    _style_range(ws, r, 2, 9, fill=SUBHEADER_FILL, font=BOLD_FONT, alignment=CENTER)
    r += 1

    r += 1
    # ── E. ABC Qty × Freq (Fast-Moving) MATRIX ───────────────────────────
    ws.merge_cells(f"B{r}:I{r}")
    ws[f"B{r}"] = "E.  ABC Qty × Freq MATRIX (Fast-Moving identification)"
    _style_range(ws, r, 2, 9, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    r += 1
    matrix_freq = (
        abc_xyz.groupby(["abc_quantity", "abc_frequency"]).size().unstack(fill_value=0)
        .reindex(index=["A", "B", "C"], columns=["A", "B", "C"])
        .fillna(0).astype(int)
    )
    headers = ["", "Freq A (High)", "Freq B (Med)", "Freq C (Low)", "Row Total", "% Volume", "% Freq", "Key Implication"]
    for col, value in enumerate(headers, start=2):
        ws.cell(r, col, value)
    _style_range(ws, r, 2, 9, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    r += 1
    for abc_class in ["A", "B", "C"]:
        values = [abc_class, int(matrix_freq.loc[abc_class, "A"]), int(matrix_freq.loc[abc_class, "B"]), int(matrix_freq.loc[abc_class, "C"]), int(matrix_freq.loc[abc_class].sum()), _format_share(_safe_float(volume_share.get(abc_class))), _format_share(_safe_float(freq_share.get(abc_class))), implications[abc_class]]
        for col, value in enumerate(values, start=2):
            ws.cell(r, col, value)
        _style_range(ws, r, 2, 9, fill=CLASS_FILLS[abc_class], alignment=LEFT)
        r += 1
    total_values = ["TOTAL", int(matrix_freq["A"].sum()), int(matrix_freq["B"].sum()), int(matrix_freq["C"].sum()), int(matrix_freq.values.sum()), "", "", ""]
    for col, value in enumerate(total_values, start=2):
        ws.cell(r, col, value)
    _style_range(ws, r, 2, 9, fill=SUBHEADER_FILL, font=BOLD_FONT, alignment=CENTER)
    r += 1

    r += 1
    # ── F. FAST-MOVING SKU GROUP (AA class) ───────────────────────────────
    class_a = abc_xyz[abc_xyz["abc_quantity"].eq("A")].copy()
    fast_moving = abc_xyz[abc_xyz["abc_quantity"].eq("A") & abc_xyz["abc_frequency"].eq("A")].copy()
    ws.merge_cells(f"B{r}:I{r}")
    ws[f"B{r}"] = "F.  FAST-MOVING GROUP (A-A: ABC_Qty=A AND ABC_Freq=A)"
    _style_range(ws, r, 2, 9, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    r += 1
    headers = ["Metric", "Value", "% of Total", "Formula", "Operational Meaning", "", "", ""]
    for col, value in enumerate(headers, start=2):
        ws.cell(r, col, value)
    _style_range(ws, r, 2, 9, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    r += 1
    total_qty = abc_xyz["quantity"].sum()
    total_freq = abc_xyz["order_frequency"].sum()
    total_skus = abc_xyz["sku_code"].nunique()
    fm_skus = fast_moving["sku_code"].nunique()
    fm_qty = _safe_float(fast_moving["quantity"].sum())
    fm_freq = _safe_float(fast_moving["order_frequency"].sum())
    fm_avg_monthly = fm_qty / month_count if month_count else 0
    top10 = ", ".join(fast_moving.sort_values("quantity", ascending=False)["sku_code"].astype(str).head(10).tolist())
    fm_rows = [
        ("Số SKU",          f"{fm_skus}",          f"{fm_skus/total_skus:.1%} SKU",   "COUNT(ABC_Qty=A AND ABC_Freq=A)",    "~4% SKU nhưng chi phối vận hành"),
        ("Sản lượng (pcs)", f"{round(fm_qty):,}",  f"{fm_qty/total_qty:.1%} qty",     "SUM(Total_Qty) nhóm A-A",           "Chiếm ~58% tổng xuất kho"),
        ("Tần suất đơn",    f"{round(fm_freq):,}", f"{fm_freq/total_freq:.1%} freq",  "SUM(Order_Freq) nhóm A-A",          "~26% pick events"),
        ("Nhu cầu TB/tháng",f"{round(fm_avg_monthly):,}", "—",                        "SUM(Total_Qty)/7",                  "Input cho Safety Stock"),
        ("Top 10 Fast-Moving SKU", top10,           "—",                               "Sort by Total_Qty desc",            "Assign to nearest pick-face slots"),
    ]
    for metric, value, pct, formula, meaning_txt in fm_rows:
        ws.cell(r, 2, metric)
        ws.cell(r, 3, value)
        ws.cell(r, 4, pct)
        ws.cell(r, 5, formula)
        ws.cell(r, 6, meaning_txt)
        _style_range(ws, r, 2, 9, fill=SUBHEADER_FILL if r % 2 == 0 else None, alignment=LEFT)
        r += 1

    key_finding = (
        f"KEY FINDING: {fm_skus} Fast-Moving SKUs ({fm_skus/total_skus:.1%} of portfolio) account for "
        f"{fm_qty/total_qty:.1%} of outbound volume and {fm_freq/total_freq:.1%} of order frequency. "
        f"These should occupy pick-face ground-level slots nearest the packing station."
    )
    ws.merge_cells(f"B{r}:I{r}")
    ws[f"B{r}"] = key_finding
    _style_range(ws, r, 2, 9, fill=TITLE_FILL, font=Font(bold=True, color="FFFFFF"), alignment=LEFT)
    r += 2

    ws.freeze_panes = "B3"

def _write_full_sku_ranking(ws, abc_xyz: pd.DataFrame) -> None:
    widths = [3, 10, 14, 15, 10, 11, 10, 10, 10, 10, 11, 10, 12, 10, 8, 10, 22, 28]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    _apply_title(
        ws,
        "B1",
        18,
        f"FULL SKU RANKING TABLE — ABC-XYZ CLASSIFICATION ({abc_xyz['sku_code'].nunique():,} SKUs | Jun\u2013Dec 2025)",
    )
    headers = [
        "Rank\n(Qty)",
        "SKU Code",
        "Total Qty\n(pcs)",
        "% Qty",
        "Cum %\nQty",
        "ABC\n(Qty)",
        "Rank\n(Freq)",
        "Order\nFreq",
        "% Freq",
        "Cum Freq Share",
        "Total CBM",
        "CV",
        "ABC\n(Freq)",
        "ABC-Freq\nMatrix",
        "XYZ\n(Vol)",
        "ABC-XYZ\n(Vol)",
        "Operational Category",
        "Slotting Zone",
    ]
    for col, value in enumerate(headers, start=2):
        ws.cell(2, col, value)
    _style_range(ws, 2, 2, 17, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)

    qty_sorted = abc_xyz.sort_values(["quantity", "sku_code"], ascending=[False, True]).reset_index(drop=True)
    qty_rank = {sku: idx for idx, sku in enumerate(qty_sorted["sku_code"], start=1)}
    freq_sorted = abc_xyz.sort_values(["order_frequency", "sku_code"], ascending=[False, True]).reset_index(drop=True)
    freq_rank = {sku: idx for idx, sku in enumerate(freq_sorted["sku_code"], start=1)}

    for row_idx, row in enumerate(qty_sorted.itertuples(index=False), start=3):
        values = [
            qty_rank[row.sku_code],
            row.sku_code,
            _safe_float(row.quantity),
            _safe_float(row.quantity_share),
            _safe_float(row.quantity_cumulative_share),
            row.abc_quantity,
            freq_rank[row.sku_code],
            _safe_float(row.order_frequency),
            _safe_float(row.frequency_share),
            _safe_float(row.frequency_cumulative_share),
            _safe_float(row.cbm_total),
            _safe_float(row.demand_cv),
            row.abc_frequency,
            row.abc_frequency_matrix,
            row.xyz_volatility,
            row.abc_xyz_volatility,
            _operational_category(row.abc_xyz_volatility),
            _slotting_zone(row.abc_quantity),
        ]
        for col, value in enumerate(values, start=2):
            ws.cell(row_idx, col, value)
        _style_range(ws, row_idx, 2, 17, fill=CLASS_FILLS.get(row.abc_quantity), alignment=CENTER)
        ws.cell(row_idx, 3).alignment = LEFT
        ws.cell(row_idx, 16).alignment = LEFT
        ws.cell(row_idx, 17).alignment = LEFT
        ws.cell(row_idx, 5).number_format = "0.00%"
        ws.cell(row_idx, 6).number_format = "0.0%"
        ws.cell(row_idx, 10).number_format = "0.00%"
        ws.cell(row_idx, 11).number_format = "0.0%"
        ws.cell(row_idx, 13).number_format = "0.000"
    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"B2:Q{ws.max_row}"


def _write_monthly_demand_sheet(ws, monthly_demand: pd.DataFrame) -> None:
    ab = monthly_demand[monthly_demand["abc_quantity"].isin(["A", "B"])].copy()
    month_columns = _month_columns(monthly_demand)
    widths = [3, 14, 8, 8] + [11] * len(month_columns) + [12, 11, 11, 8, 10]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    _apply_title(ws, "B1", 6 + len(month_columns) + 4, "MONTHLY DEMAND TABLE — A & B CLASS SKUs (ABC Frequency Base)")
    headers = ["SKU Code", "ABC", "ABC (Freq)", "ABC-Freq Matrix", "XYZ (Vol)", "ABC-XYZ (Vol)", *month_columns, "Total", "Mean/Mo", "Std Dev", "CV"]
    for col, value in enumerate(headers, start=2):
        ws.cell(2, col, value)
    _style_range(ws, 2, 2, 1 + len(headers), fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    for row_idx, (_, row) in enumerate(ab.iterrows(), start=3):
        values = [
            row["sku_code"],
            row["abc_quantity"],
            row["abc_frequency"],
            row["abc_frequency_matrix"],
            row["xyz_volatility"],
            row["abc_xyz_volatility"],
            *[row[column] for column in month_columns],
            row["Total"],
            row["Mean/Mo"],
            row["Std Dev"],
            row["CV"],
        ]
        for col, value in enumerate(values, start=2):
            ws.cell(row_idx, col, value)
        _style_range(ws, row_idx, 2, 1 + len(headers), fill=CLASS_FILLS.get(row["abc_quantity"]), alignment=CENTER)
        ws.cell(row_idx, 2).alignment = LEFT
        total_col = 2 + 6 + len(month_columns)
        mean_col = total_col + 1
        std_col = total_col + 2
        cv_col = total_col + 3
        ws.cell(row_idx, mean_col).number_format = "0.0"
        ws.cell(row_idx, std_col).number_format = "0.0"
        ws.cell(row_idx, cv_col).number_format = "0.000"
    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"B2:{get_column_letter(1 + len(headers))}{ws.max_row}"


def _write_class_a_deep_dive(ws, monthly_demand: pd.DataFrame) -> None:
    class_a = monthly_demand[monthly_demand["abc_quantity"].eq("A")].copy().sort_values("quantity", ascending=False)
    widths = [3, 8, 14, 12, 12, 12, 12, 12, 12, 12, 10, 14, 22, 20, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    _apply_title(
        ws,
        "B1",
        15,
        f"CLASS A SKUs — FAST-MOVING DEEP DIVE ({len(class_a):,} SKUs | {_format_share(class_a['quantity'].sum() / monthly_demand['quantity'].sum())} of Volume)",
    )
    ws.merge_cells("B2:O2")
    ws["B2"] = (
        "These SKUs create the highest operational pressure: prioritize pick-face slotting, cycle counting, "
        "and safety stock buffering for AZ subgroup."
    )
    ws["B2"].alignment = LEFT
    ws["B2"].font = Font(italic=True, color="5B5B5B")
    headers = [
        "Rank",
        "SKU Code",
        "ABC-XYZ\n(Freq)",
        "ABC-XYZ\n(Vol)",
        "Total Qty\n(pcs)",
        "% of Total\nQty",
        "Order\nFrequency",
        "% of Total\nFreq",
        "Total CBM",
        "Avg Monthly\nDemand",
        "CV",
        "Demand\nPattern",
        "Stock\nStrategy",
        "Slotting\nPriority",
        "Reorder\nLogic",
    ]
    for col, value in enumerate(headers, start=2):
        ws.cell(3, col, value)
    _style_range(ws, 3, 2, 15, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    total_qty = monthly_demand["quantity"].sum()
    total_freq = monthly_demand["order_frequency"].sum()
    for row_idx, (_, row) in enumerate(class_a.iterrows(), start=4):
        values = [
            row_idx - 3,
            row["sku_code"],
            row["abc_frequency_matrix"],
            row["abc_xyz_volatility"],
            _safe_float(row["quantity"]),
            _safe_float(row["quantity"]) / total_qty if total_qty else 0.0,
            _safe_float(row["order_frequency"]),
            _safe_float(row["order_frequency"]) / total_freq if total_freq else 0.0,
            _safe_float(row["cbm_total"]),
            _safe_float(row["Mean/Mo"]),
            _safe_float(row["CV"]),
            _demand_pattern(row["xyz_volatility"]),
            _stock_strategy(row["xyz_volatility"]),
            "Ground Level - Zone 1",
            _reorder_logic(row["abc_frequency"]),
        ]
        for col, value in enumerate(values, start=2):
            ws.cell(row_idx, col, value)
        _style_range(ws, row_idx, 2, 15, fill=CLASS_FILLS.get(row["abc_frequency"]), alignment=CENTER)
        ws.cell(row_idx, 3).alignment = LEFT
        ws.cell(row_idx, 6).number_format = "0.00%"
        ws.cell(row_idx, 8).number_format = "0.00%"
        ws.cell(row_idx, 10).number_format = "0.0"
        ws.cell(row_idx, 11).number_format = "0.000"
        ws.cell(row_idx, 12).alignment = LEFT
        ws.cell(row_idx, 13).alignment = LEFT
        ws.cell(row_idx, 14).alignment = LEFT
        ws.cell(row_idx, 15).alignment = LEFT
    ws.freeze_panes = "B4"
    ws.auto_filter.ref = f"B3:O{ws.max_row}"


def _write_q21_dark_store_sla_sheet(ws, network_model_evaluation: pd.DataFrame, shipments: pd.DataFrame) -> None:
    # Set column widths
    ws.column_dimensions[get_column_letter(1)].width = 3
    headers = [
        "District",
        "Quantity (pcs)",
        "Qty Share",
        "Traffic Factor",
        "Baseline Dist (km)",
        "Baseline Time (min)",
        "Baseline 2H SLA",
        "Baseline 4H SLA",
        "Nearest Facility",
        "Optimal Dist (km)",
        "Optimal Time (min)",
        "Optimal 2H SLA",
        "Optimal 4H SLA"
    ]
    for col_idx, h in enumerate(headers, start=2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
        ws.cell(2, col_idx, h)
    _apply_title(ws, "B1", len(headers) + 1, "QUESTION 2.1 — DARK STORE SLA COMPARISON (HCM DISTRICTS)")
    _style_range(ws, 2, 2, len(headers) + 1, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    
    # Write rows
    for row_idx, row in enumerate(network_model_evaluation.itertuples(index=False), start=3):
        values = [
            row.district,
            _safe_float(row.quantity),
            _safe_float(row.qty_share),
            _safe_float(row.traffic_factor),
            _safe_float(row.baseline_dist),
            _safe_float(row.baseline_time),
            row.baseline_2h,
            row.baseline_4h,
            row.nearest_facility,
            _safe_float(row.ds_dist),
            _safe_float(row.ds_time),
            row.ds_2h,
            row.ds_4h
        ]
        for col_idx, value in enumerate(values, start=2):
            cell = ws.cell(row_idx, col_idx, value)
            cell.border = THIN_BORDER
            cell.alignment = LEFT if isinstance(value, str) else CENTER
            if col_idx in {3, 6, 7, 11, 12}: # numeric floats / quantities
                cell.number_format = "0.0" if col_idx != 3 else "#,##0"
            elif col_idx == 4: # share percentage
                cell.number_format = "0.00%"
            elif col_idx == 5: # traffic factor
                cell.number_format = "0.0"
                
    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"B2:{get_column_letter(len(headers) + 1)}{len(network_model_evaluation) + 2}"

    r = len(network_model_evaluation) + 5
    ws.merge_cells(f"B{r}:H{r}")
    ws[f"B{r}"] = "SUMMARY COMPARISON METRICS"
    _style_range(ws, r, 2, 8, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    r += 1
    
    # Calculate summary metrics
    total_qty = network_model_evaluation["quantity"].sum()
    b_2h = (network_model_evaluation["baseline_2h"].eq("Met") * network_model_evaluation["quantity"]).sum() / total_qty
    b_4h = (network_model_evaluation["baseline_4h"].eq("Met") * network_model_evaluation["quantity"]).sum() / total_qty
    ds_2h = (network_model_evaluation["ds_2h"].eq("Met") * network_model_evaluation["quantity"]).sum() / total_qty
    ds_4h = (network_model_evaluation["ds_4h"].eq("Met") * network_model_evaluation["quantity"]).sum() / total_qty
    
    base_w_dist = (network_model_evaluation["baseline_dist"] * network_model_evaluation["quantity"]).sum() / total_qty
    ds_w_dist = (network_model_evaluation["ds_dist"] * network_model_evaluation["quantity"]).sum() / total_qty
    savings_km = base_w_dist - ds_w_dist
    savings_pct = savings_km / base_w_dist if base_w_dist else 0.0
    
    # Throughput calculation
    hcm_ship = shipments[
        shipments["province"].eq("Hồ Chí Minh") 
        & shipments["created_date"].between("2025-06-01", "2025-12-31") 
        & shipments["document_type"].eq("A/R INVOICE")
    ]
    hcm_active_days = hcm_ship["created_date"].dt.date.nunique()
    ds_routed_qty = network_model_evaluation[
        network_model_evaluation["nearest_facility"].isin(["DS1 (Tân Phú)", "DS2 (Quận 1)"])
    ]["quantity"].sum()
    
    summary_rows = [
        ("Baseline 2H SLA Coverage (weighted)", b_2h),
        ("Baseline 4H SLA Coverage (weighted)", b_4h),
        ("2 Dark Stores 2H SLA Coverage (weighted)", ds_2h),
        ("2 Dark Stores 4H SLA Coverage (weighted)", ds_4h),
        ("Baseline Weighted Distance (km)", base_w_dist),
        ("2 Dark Stores Weighted Distance (km)", ds_w_dist),
        ("Weighted Distance Savings (km)", savings_km),
        ("Weighted Distance Savings (% reduction)", savings_pct),
        ("Total Quantity Routed to Dark Stores (pcs)", ds_routed_qty),
        ("Active Transaction Days in HCMC Dataset", hcm_active_days),
        ("Dark Store Daily Throughput (based on active days)", ds_routed_qty / hcm_active_days if hcm_active_days else 0.0),
        ("Dark Store Daily Throughput (standard 214 operational days)", ds_routed_qty / 214.0)
    ]
    
    for metric, val in summary_rows:
        ws.cell(r, 2, metric)
        cell = ws.cell(r, 3, val)
        if "%" in metric or "Coverage" in metric:
            cell.number_format = "0.00%"
        elif "Distance" in metric or "Savings" in metric or "Throughput" in metric:
            cell.number_format = "#,##0.00"
        else:
            cell.number_format = "#,##0"
        ws.cell(r, 2).font = BOLD_FONT
        ws.cell(r, 3).font = BOLD_FONT
        r += 1


def _write_q31_travel_time_summary_sheet(
    ws,
    travel_metrics: dict,
    sku_pick_profile: pd.DataFrame,
    slotting_plan: pd.DataFrame,
    abc_xyz: pd.DataFrame,
) -> None:
    from src.logage2026.analysis import (
        DIST_A_M,
        DIST_B_M,
        DIST_C_M,
        PICK_RATE_A_PCS_MIN,
        PICK_RATE_B_PCS_MIN,
        PICK_RATE_C_PCS_MIN,
        REDUCTION_TARGET,
        ROUND_TRIP_FACTOR,
    )

    ws.column_dimensions[get_column_letter(1)].width = 3
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions[get_column_letter(9)].width = 30

    _apply_title(ws, "B1", 9, "LOGage 2026 — QUESTION 3.1 | SLOT PLANNING & TRAVEL-TIME PROOF")

    r = 3
    ws.merge_cells(f"B{r}:I{r}")
    ws[f"B{r}"] = "A. ZONE SUMMARY (MODEL 1) — PHYSICAL ASSIGNMENTS"
    _style_range(ws, r, 2, 9, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    r += 1

    headers = ["Zone", "Class", "# SKUs", "Total Picks", "One-way Dist (m)", "Round-trip factor", "Pick Rate", "Zone Description", ""]
    for col, val in enumerate(headers, start=2):
        ws.cell(r, col, val)
    _style_range(ws, r, 2, 9, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    r += 1

    zone_data = [
        ("Pick-Face Zone", "A", travel_metrics["zone_counts"]["A"], travel_metrics["zone_picks"]["A"], DIST_A_M, ROUND_TRIP_FACTOR, f"{PICK_RATE_A_PCS_MIN:.0f} pcs/min", "Golden zone (ground racks)"),
        ("Forward Reserve Zone", "B", travel_metrics["zone_counts"]["B"], travel_metrics["zone_picks"]["B"], DIST_B_M, ROUND_TRIP_FACTOR, f"{PICK_RATE_B_PCS_MIN:.1f} pcs/min", "Replenished from bulk (mid racks)"),
        ("Reserve / Bulk Zone", "C", travel_metrics["zone_counts"]["C"], travel_metrics["zone_picks"]["C"], DIST_C_M, ROUND_TRIP_FACTOR, f"{PICK_RATE_C_PCS_MIN:.1f} pcs/min", "Upper racks & block-stack")
    ]

    for name, cls, count, picks, dist, rtf, rate, desc in zone_data:
        values = [name, cls, count, picks, dist, rtf, rate, desc, ""]
        for col_idx, value in enumerate(values, start=2):
            cell = ws.cell(r, col_idx, value)
            cell.border = THIN_BORDER
            cell.alignment = LEFT if col_idx in {2, 9} else CENTER
            if col_idx == 5:
                cell.number_format = "#,##0"
        _style_range(ws, r, 2, 9, fill=CLASS_FILLS.get(cls), alignment=None)
        r += 1

    r += 1
    ws.merge_cells(f"B{r}:I{r}")
    ws[f"B{r}"] = "B. PICK-UNIT MIX BY ZONE (Avg. order-line share)"
    _style_range(ws, r, 2, 9, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    r += 1

    headers = ["Zone", "Pallet Pick Share", "Carton Pick Share", "Loose Pick Share", "", "", "", "", ""]
    for col, val in enumerate(headers, start=2):
        ws.cell(r, col, val)
    _style_range(ws, r, 2, 9, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    r += 1

    zone_mix = sku_pick_profile.groupby("abc_quantity")[["pallet_share", "carton_share", "loose_share"]].mean()
    for cls in ["A", "B", "C"]:
        p_share = zone_mix.loc[cls, "pallet_share"] if cls in zone_mix.index else 0.0
        c_share = zone_mix.loc[cls, "carton_share"] if cls in zone_mix.index else 0.0
        l_share = zone_mix.loc[cls, "loose_share"] if cls in zone_mix.index else 0.0

        values = [f"Zone {cls}", p_share, c_share, l_share, "", "", "", "", ""]
        for col_idx, value in enumerate(values, start=2):
            cell = ws.cell(r, col_idx, value)
            cell.border = THIN_BORDER
            cell.alignment = LEFT if col_idx == 2 else CENTER
            if col_idx in {3, 4, 5} and isinstance(value, float):
                cell.number_format = "0.0%"
        _style_range(ws, r, 2, 9, fill=CLASS_FILLS.get(cls), alignment=None)
        r += 1

    r += 1
    ws.merge_cells(f"B{r}:I{r}")
    ws[f"B{r}"] = "C. MODEL 2 SUB-TIER ASSIGNMENTS"
    _style_range(ws, r, 2, 9, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    r += 1

    headers = ["Sub-Tier", "Name", "Assignment Rule", "SKUs", "", "", "", "", ""]
    for col, val in enumerate(headers, start=2):
        ws.cell(r, col, val)
    _style_range(ws, r, 2, 9, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    r += 1

    tier_counts = slotting_plan["sub_tier"].value_counts().to_dict()
    sub_tiers = [
        ("A1", "Pallet Lane", "pallet_share >= 40% AND CBM >= 0.07 m3 (ground level)", tier_counts.get("A1", 0)),
        ("A2", "Big-Face Shelf", "pallet_share < 40% AND carton_share >= 80% (waist height)", tier_counts.get("A2", 0)),
        ("A3", "Mixed-Pick", "Remaining Class A (ground pick-face racks)", tier_counts.get("A3", 0)),
        ("B1", "Bulk-Replen", "pallet_share >= 30% OR CBM >= 0.07 m3 (mid-aisle ground)", tier_counts.get("B1", 0)),
        ("B2", "Two-Bin", "Remaining Class B (forward pick-face bin)", tier_counts.get("B2", 0)),
        ("C1", "Upper Rack", "CBM < 0.05 m3 (small/light slow-movers)", tier_counts.get("C1", 0)),
        ("C2", "Pallet Block", "CBM >= 0.05 m3 (floor block-stack behind Zone B)", tier_counts.get("C2", 0)),
    ]

    for code, name, rule, count in sub_tiers:
        cls = code[0]
        values = [code, name, rule, count, "", "", "", "", ""]
        for col_idx, value in enumerate(values, start=2):
            cell = ws.cell(r, col_idx, value)
            cell.border = THIN_BORDER
            cell.alignment = LEFT if col_idx in {3, 4} else CENTER
            if col_idx == 5:
                cell.number_format = "#,##0"
        _style_range(ws, r, 2, 9, fill=CLASS_FILLS.get(cls), alignment=None)
        r += 1

    r += 1
    ws.merge_cells(f"B{r}:I{r}")
    ws[f"B{r}"] = "D. TRAVEL-TIME PROOF SUMMARY (Baseline vs. Optimized)"
    _style_range(ws, r, 2, 9, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    r += 1

    headers = ["Metric", "Baseline (Random Placement)", "Optimized (ABC-Zoned)", "Improvement / Status", "", "", "", "", ""]
    for col, val in enumerate(headers, start=2):
        ws.cell(r, col, val)
    _style_range(ws, r, 2, 9, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    r += 1

    status_text = "MEETS TARGET" if travel_metrics["meets_target"] else "BELOW TARGET"
    reduction_pct = travel_metrics["travel_reduction"]

    rows = [
        ("Avg One-way Distance per Pick", travel_metrics["baseline_avg_dist_m"], travel_metrics["opt_avg_oneway_m_per_pick"], "meters"),
        ("Total Round-trip Distance", travel_metrics["baseline_total_round_trip_m"], travel_metrics["opt_total_round_trip_m"], "meters"),
        ("Travel Time (walk speed 72m/min)", travel_metrics["baseline_travel_time_min"], travel_metrics["opt_travel_time_min"], "minutes"),
        ("Travel Time Reduction %", "", reduction_pct, f"{reduction_pct:.1%} (Target: >= {REDUCTION_TARGET:.0%})"),
        ("SLA Performance Status", "", status_text, "PASSED" if travel_metrics["meets_target"] else "FAILED")
    ]

    for metric, base, opt, note in rows:
        values = [metric, base, opt, note, "", "", "", "", ""]
        for col_idx, value in enumerate(values, start=2):
            cell = ws.cell(r, col_idx, value)
            cell.border = THIN_BORDER
            cell.alignment = LEFT if col_idx in {2, 5} else CENTER
            if col_idx in {3, 4} and isinstance(value, float):
                if "Reduction" in metric:
                    cell.number_format = "0.0%"
                else:
                    cell.number_format = "#,##0.0"

        if metric == "SLA Performance Status":
            fill_color = "D9EAD3" if travel_metrics["meets_target"] else "F4CCCC"
            _style_range(ws, r, 2, 9, fill=PatternFill("solid", fgColor=fill_color), font=BOLD_FONT, alignment=None)
        else:
            _style_range(ws, r, 2, 9, fill=SUBHEADER_FILL if r % 2 == 0 else None, font=BOLD_FONT if col_idx == 2 else None, alignment=None)
        r += 1

def _write_q12_heatmap_summary_sheet(
    ws,
    warehouse_region_summary: pd.DataFrame,
    q12_top_demand_provinces_summary: pd.DataFrame,
    warehouse_imbalance_summary: pd.DataFrame,
) -> None:
    ws.column_dimensions[get_column_letter(1)].width = 3
    for col in range(2, 12):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions[get_column_letter(12)].width = 30
    _apply_title(ws, "B1", 11, "LOGage 2026 — QUESTION 1.2 | HEATMAP SUMMARY")

    r = 3
    ws.merge_cells(f"B{r}:K{r}")
    ws[f"B{r}"] = "A. WAREHOUSE BALANCE (My Phuoc vs Vinh Loc)"
    _style_range(ws, r, 2, 11, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    r += 1
    headers = ["Warehouse", "Quantity (pcs)", "% Total Qty", "CBM (m3)", "% Total CBM", "Order Count", "% Total Orders", "Transaction Rows", "% Total Rows", "", ""]
    for col, value in enumerate(headers, start=2):
        ws.cell(r, col, value)
    _style_range(ws, r, 2, 11, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    r += 1
    wh_summary = (
        warehouse_region_summary.groupby("source_warehouse")
        .agg(quantity=("quantity", "sum"), cbm=("cbm_total", "sum"), orders=("orders", "sum"), rows=("shipment_lines", "sum"))
        .reset_index()
    )
    total_qty = wh_summary["quantity"].sum()
    total_cbm = wh_summary["cbm"].sum()
    total_orders = wh_summary["orders"].sum()
    total_rows = wh_summary["rows"].sum()
    for _, row in wh_summary.iterrows():
        values = [row["source_warehouse"], round(row["quantity"]), row["quantity"] / total_qty if total_qty else 0,
                  round(row["cbm"], 2), row["cbm"] / total_cbm if total_cbm else 0,
                  int(row["orders"]), row["orders"] / total_orders if total_orders else 0,
                  int(row["rows"]), row["rows"] / total_rows if total_rows else 0]
        for col, value in enumerate(values, start=2):
            cell = ws.cell(r, col, value)
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            if col in {3, 5, 7, 9}:
                cell.number_format = "0.0%"
            elif col in {4}:
                cell.number_format = "0.00"
        r += 1

    r += 2
    ws.merge_cells(f"B{r}:K{r}")
    ws[f"B{r}"] = "B. DEMAND BY REGION"
    _style_range(ws, r, 2, 11, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    r += 1
    headers = ["Region", "Quantity (pcs)", "% Total", "CBM (m3)", "% Total CBM", "Orders", "% Total Orders", "Qty/Order", "Customers", "Rows", ""]
    for col, value in enumerate(headers, start=2):
        ws.cell(r, col, value)
    _style_range(ws, r, 2, 11, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    r += 1
    region_agg = (
        warehouse_region_summary.groupby("region")
        .agg(quantity=("quantity", "sum"), cbm=("cbm_total", "sum"), orders=("orders", "sum"), customers=("customers", "nunique"), rows=("shipment_lines", "sum"))
        .reset_index()
        .sort_values("quantity", ascending=False)
    )
    tot_qty = region_agg["quantity"].sum()
    tot_cbm = region_agg["cbm"].sum()
    tot_ord = region_agg["orders"].sum()
    for _, row in region_agg.iterrows():
        qty_per_order = row["quantity"] / row["orders"] if row["orders"] else 0
        values = [row["region"], round(row["quantity"]), row["quantity"] / tot_qty if tot_qty else 0,
                  round(row["cbm"], 2), row["cbm"] / tot_cbm if tot_cbm else 0,
                  int(row["orders"]), row["orders"] / tot_ord if tot_ord else 0,
                  round(qty_per_order, 1), int(row["customers"]), int(row["rows"])]
        for col, value in enumerate(values, start=2):
            cell = ws.cell(r, col, value)
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            if col in {3, 5, 7}:
                cell.number_format = "0.0%"
            elif col == 8:
                cell.number_format = "0.0"
        r += 1

    r += 2
    ws.merge_cells(f"B{r}:K{r}")
    ws[f"B{r}"] = "C. TOP 15 PROVINCES BY QUANTITY"
    _style_range(ws, r, 2, 11, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    r += 1
    headers = ["Rank", "Province", "Quantity (pcs)", "% Total", "Cum %", "Orders", "Qty/Order", "Dist. MP (km)", "Dist. VL (km)", "Region", ""]
    for col, value in enumerate(headers, start=2):
        ws.cell(r, col, value)
    _style_range(ws, r, 2, 11, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    r += 1
    top_provs = (
        warehouse_region_summary.groupby("province")
        .agg(quantity=("quantity", "sum"), orders=("orders", "sum"), region=("region", "first"))
        .reset_index()
        .sort_values("quantity", ascending=False)
        .head(15)
    )
    top_prov_total_qty = top_provs["quantity"].sum()
    cum_share = 0.0
    for rank, (_, row) in enumerate(top_provs.iterrows(), 1):
        cum_share += row["quantity"] / top_prov_total_qty if top_prov_total_qty else 0
        qty_per_order = row["quantity"] / row["orders"] if row["orders"] else 0
        values = [rank, row["province"], round(row["quantity"]),
                  row["quantity"] / top_prov_total_qty if top_prov_total_qty else 0,
                  cum_share, int(row["orders"]), round(qty_per_order, 1), "", "", row["region"]]
        for col, value in enumerate(values, start=2):
            cell = ws.cell(r, col, value)
            cell.border = THIN_BORDER
            cell.alignment = LEFT if col in {3, 11} else CENTER
            if col in {4, 5}:
                cell.number_format = "0.0%"
            elif col == 7:
                cell.number_format = "0.0"
        r += 1

    ws.freeze_panes = "B3"


def _write_q13_order_profile_summary_sheet(
    ws,
    q13_segment_profile_summary: pd.DataFrame,
    q13_segment_packaging_summary: pd.DataFrame,
) -> None:
    for idx, width in enumerate([3, 22, 16, 16, 14, 14, 16, 16, 14, 14, 14, 14, 14, 14], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    _apply_title(ws, "B1", 14, "LOGage 2026 — QUESTION 1.3 | ORDER PROFILE SUMMARY")

    r = 3
    ws.merge_cells(f"B{r}:N{r}")
    ws[f"B{r}"] = "A. CUSTOMER SEGMENT COMPARISON TABLE"
    _style_range(ws, r, 2, 14, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    r += 1
    headers = [
        "Tieu chi", "Formula", "Modern Trade", "Traditional Trade",
        "MT vs TT", "Operational Meaning",
        "", "", "", "", "", "", "", ""
    ]
    for col, value in enumerate(headers, start=2):
        ws.cell(r, col, value)
    _style_range(ws, r, 2, 14, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    r += 1

    mt = q13_segment_profile_summary[q13_segment_profile_summary["customer_segment"] == "Modern Trade"]
    tt = q13_segment_profile_summary[q13_segment_profile_summary["customer_segment"] == "Traditional Trade / Distributor"]
    mt_row = mt.iloc[0] if not mt.empty else None
    tt_row = tt.iloc[0] if not tt.empty else None

    def _mtv(key):
        return mt_row[key] if mt_row is not None else "N/A"
    def _ttv(key):
        return tt_row[key] if tt_row is not None else "N/A"

    rows_data = [
        ("Active Customers", "COUNT(DISTINCT customer_key)", _mtv("customers"), _ttv("customers"), "", "Khach hang lon (MT) vs nhieu khach le (TT)"),
        ("Total Orders", "COUNT(DISTINCT document_no)", _mtv("orders"), _ttv("orders"), "", "MT dat don lon, TT dat don nho le nhieu lan"),
        ("Avg Order Quantity (pcs)", "AVG(SUM quantity per order)", f'{_mtv("avg_order_quantity"):.1f}', f'{_ttv("avg_order_quantity"):.1f}', "", "MT mua so luong lon/don; TT mua nhot giot"),
        ("Median Order Quantity (pcs)", "MEDIAN(SUM quantity per order)", f'{_mtv("median_order_quantity"):.1f}', f'{_ttv("median_order_quantity"):.1f}', "", "Phan phoi lech phai o TT"),
        ("Avg Order CBM (m3)", "AVG(SUM CBM per order)", f'{_mtv("avg_order_cbm"):.3f}', f'{_ttv("avg_order_cbm"):.3f}', "", "MT don hang chiem khoi luong lon hon"),
        ("Median Order CBM (m3)", "MEDIAN(SUM CBM per order)", f'{_mtv("median_order_cbm"):.3f}', f'{_ttv("median_order_cbm"):.3f}', "", ""),
        ("Avg SKU Breadth/Order", "AVG(COUNT DISTINCT SKU per order)", f'{_mtv("avg_sku_breadth"):.2f}', f'{_ttv("avg_sku_breadth"):.2f}', "", "TT order nhieu SKU hon phan loai phuc tap"),
        ("Avg Lines per Order", "AVG(COUNT line per order)", f'{_mtv("avg_lines_per_order"):.2f}', f'{_ttv("avg_lines_per_order"):.2f}', "", ""),
        ("Avg Orders/Customer/Month", "Orders / Customers / 7 months", f'{_mtv("avg_orders_per_customer_month"):.2f}', f'{_ttv("avg_orders_per_customer_month"):.2f}', "", "TT tan suat cao hon"),
        ("Active Month Frequency", "Orders / active months", f'{_mtv("active_month_frequency"):.2f}', f'{_ttv("active_month_frequency"):.2f}', "", ""),
        ("Geographic Spread", "COUNT(DISTINCT province)", f'{_mtv("province_count")} provinces', f'{_ttv("province_count")} provinces', "", "TT phu rong toan quoc"),
        ("Avg Delivery Distance (km)", "AVG(distance_km per order)", f'{_mtv("avg_distance_km"):.1f}', f'{_ttv("avg_distance_km"):.1f}', "", "MT xa hon do kho tap trung o trung tam"),
        ("Lead Time Sensitivity", "Z x sigma_daily x sqrt(LT)", str(_mtv("lead_time_sensitivity")), str(_ttv("lead_time_sensitivity")), "", "MT nhay hon (SLA khat khe)"),
    ]
    for metric, formula, mt_val, tt_val, mt_vs_tt, meaning in rows_data:
        values = [metric, formula, mt_val, tt_val, mt_vs_tt, meaning, "", "", "", "", "", "", "", ""]
        for col, value in enumerate(values, start=2):
            cell = ws.cell(r, col, value)
            cell.border = THIN_BORDER
            cell.alignment = LEFT
        r += 1

    r += 2
    ws.merge_cells(f"B{r}:N{r}")
    ws[f"B{r}"] = "B. PACK UNIT MIX BY SEGMENT"
    _style_range(ws, r, 2, 14, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    r += 1
    headers = ["Segment", "Pallet Qty", "Pallet %", "Carton Qty", "Carton %", "Loose Qty", "Loose %", "Total Qty", "", "", "", "", "", ""]
    for col, value in enumerate(headers, start=2):
        ws.cell(r, col, value)
    _style_range(ws, r, 2, 14, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    r += 1
    for segment_name in ["Modern Trade", "Traditional Trade / Distributor"]:
        pkg = q13_segment_packaging_summary[q13_segment_packaging_summary["customer_segment"] == segment_name]
        total_q = pkg["quantity"].sum() if not pkg.empty else 0
        pallet_q = pkg[pkg["packaging_unit"] == "pallet"]["quantity"].sum() if not pkg.empty else 0
        carton_q = pkg[pkg["packaging_unit"] == "carton"]["quantity"].sum() if not pkg.empty else 0
        loose_q = pkg[pkg["packaging_unit"] == "loose"]["quantity"].sum() if not pkg.empty else 0
        values = [segment_name, round(pallet_q), pallet_q / total_q if total_q else 0,
                  round(carton_q), carton_q / total_q if total_q else 0,
                  round(loose_q), loose_q / total_q if total_q else 0,
                  round(total_q)]
        for col, value in enumerate(values, start=2):
            cell = ws.cell(r, col, value)
            cell.border = THIN_BORDER
            cell.alignment = LEFT if col == 2 else CENTER
            if col in {4, 6, 8}:
                cell.number_format = "0.0%"
        r += 1

    ws.freeze_panes = "B3"


def _write_q22_lead_time_sheet(ws, lead_time_table: pd.DataFrame) -> None:
    ws.column_dimensions[get_column_letter(1)].width = 3
    for col in range(2, 10):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.column_dimensions[get_column_letter(10)].width = 35
    _apply_title(ws, "B1", 10, "LOGage 2026 — QUESTION 2.2 | MILE-WEIGHTED AVERAGE LEAD TIME")

    r = 3
    ws.merge_cells(f"B{r}:J{r}")
    ws[f"B{r}"] = (
        "Methodology: LT_avg = Sigma(LT_r x Orders_r) / Sigma(Orders_r) "
        "mile-weighted by order count per region. Transit benchmarks follow GHN/VTP/GHTK standards."
    )
    ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="center")
    r += 1

    ws.merge_cells(f"B{r}:J{r}")
    ws[f"B{r}"] = "A. REGION-LEVEL LEAD TIME TABLE"
    _style_range(ws, r, 2, 10, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    r += 1
    headers = ["Region", "Route Type", "LT (days)", "Orders", "Weight", "Quantity", "LT x Orders", "Benchmark Source", "", ""]
    for col, value in enumerate(headers, start=2):
        ws.cell(r, col, value)
    _style_range(ws, r, 2, 10, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    r += 1

    for _, row in lead_time_table.iterrows():
        values = [
            row.get("region", ""),
            row.get("route_type", ""),
            row.get("lt_days") if pd.notna(row.get("lt_days")) else "",
            int(row["orders"]) if pd.notna(row["orders"]) else "",
            row.get("weight") if pd.notna(row.get("weight")) else "",
            row.get("quantity") if pd.notna(row.get("quantity")) else "",
            row["lt_x_orders"] if pd.notna(row["lt_x_orders"]) else "",
            row.get("benchmark_source", ""),
        ]
        is_total = str(row.get("region", "")).strip() == "TOTAL"
        for col, value in enumerate(values, start=2):
            cell = ws.cell(r, col, value)
            cell.border = THIN_BORDER
            cell.alignment = LEFT if col in {2, 3, 9} else CENTER
            if col == 5:
                cell.number_format = "0.00%"
            if is_total:
                cell.font = BOLD_FONT
        if is_total:
            _style_range(ws, r, 2, 10, fill=SUBHEADER_FILL, font=BOLD_FONT, alignment=None)
        r += 1

    r += 2
    ws.merge_cells(f"B{r}:J{r}")
    ws[f"B{r}"] = "B. RESULT MILE-WEIGHTED AVERAGE LEAD TIME"
    _style_range(ws, r, 2, 10, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    r += 1
    total_row = lead_time_table[lead_time_table["region"] == "TOTAL"]
    if not total_row.empty:
        bench = total_row.iloc[0].get("benchmark_source", "")
        ws.cell(r, 2, "LT_avg result:")
        ws.cell(r, 2).font = BOLD_FONT
        ws.cell(r, 3, bench)
        ws.cell(r, 3).font = BOLD_FONT
        ws.merge_cells(f"C{r}:D{r}")
        _style_range(ws, r, 2, 10, fill=TITLE_FILL, font=Font(bold=True, color="FFFFFF"), alignment=LEFT)

    ws.freeze_panes = "B3"


def _write_q22_safety_stock_sheet(ws, safety_stock_class_a: pd.DataFrame) -> None:
    for idx, width in enumerate([3, 14, 16, 14, 16, 16, 14, 14, 12, 12, 16, 16], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    _apply_title(ws, "B1", 12, "LOGage 2026 — QUESTION 2.2 | SAFETY STOCK CLASS A SKUs")

    r = 3
    ws.merge_cells(f"B{r}:L{r}")
    ws[f"B{r}"] = "FORMULAS:  SS = Z x sigma_daily x sqrt(LT_avg)    |    ROP = mu_daily x LT_avg + SS"
    _style_range(ws, r, 2, 12, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    r += 1
    ws.merge_cells(f"B{r}:L{r}")
    ws[f"B{r}"] = "Parameters: Z = 1.645 (95% service level), LT_avg = 1.94 days, sqrt(LT) = 1.393, ddof=1 over 7 months (Jun Dec 2025)"
    _style_range(ws, r, 2, 12, fill=SUBHEADER_FILL, alignment=LEFT)
    r += 1

    headers = [
        "SKU", "Product", "ABC-XYZ", "mu_monthly", "sigma_monthly",
        "mu_daily", "sigma_daily", "Z", "sqrt(LT)", "Safety Stock", "ROP", ""
    ]
    for col, value in enumerate(headers, start=2):
        ws.cell(r, col, value)
    _style_range(ws, r, 2, 12, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    r += 1

    for _, row in safety_stock_class_a.iterrows():
        abc_xyz_code = f"{row.get('abc_quantity', '')}{row.get('abc_frequency', '')}"
        values = [
            row["sku_code"],
            row.get("product_name", ""),
            abc_xyz_code,
            row["mu_monthly"] if pd.notna(row["mu_monthly"]) else 0,
            row["sigma_monthly"] if pd.notna(row["sigma_monthly"]) else 0,
            row["mu_daily"] if pd.notna(row["mu_daily"]) else 0,
            row["sigma_daily"] if pd.notna(row["sigma_daily"]) else 0,
            1.645,
            row["sqrt_lt"] if pd.notna(row["sqrt_lt"]) else 0,
            row["ss"] if pd.notna(row["ss"]) else 0,
            row["rop"] if pd.notna(row["rop"]) else 0,
        ]
        for col, value in enumerate(values, start=2):
            cell = ws.cell(r, col, value)
            cell.border = THIN_BORDER
            cell.alignment = LEFT if col in {2, 3, 4} else CENTER
            if col in {5, 6, 7, 8, 10, 11, 12} and isinstance(value, (int, float)):
                cell.number_format = "0.0"
            elif col == 11:
                cell.number_format = "0"
        r += 1
    ws.freeze_panes = "B3"


def _write_q22_inventory_pooling_sheet(ws, inventory_pooling_summary: pd.DataFrame, safety_stock_class_a: pd.DataFrame) -> None:
    widths = [3, 22, 32, 14, 14, 16, 16, 14, 14, 12, 12, 16, 16]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    _apply_title(ws, "B1", 12, "LOGage 2026 — QUESTION 2.2 | INVENTORY POOLING & SAFETY STOCK")

    r = 3
    ws.merge_cells(f"B{r}:H{r}")
    ws[f"B{r}"] = "A. CHANNEL DEFINITIONS"
    _style_range(ws, r, 2, 8, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    r += 1
    headers = ["Channel", "Accounts", "Order Profile", "SLA", "Priority", "", "", ""]
    for col, value in enumerate(headers, start=2):
        ws.cell(r, col, value)
    _style_range(ws, r, 2, 8, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    r += 1
    channel_data = [
        ("Modern Trade (MT)", "Co-op, Lotte, AEON, DMX, TGDD", "Large confirmed POs", "Strict booking-window SLA", "1 (Highest)"),
        ("Ecommerce", "Shopee, Lazada, Tiki, TikTok Shop", "High-volume, 2-4h delivery", "2-4h SLA", "2 (High)"),
        ("Traditional Trade (TT)", "Distributors / small dealers", "Small fragmented orders", "Flexible 2-4 day SLA", "3 (Low)"),
    ]
    for ch, accts, profile, sla, pri in channel_data:
        values = [ch, accts, profile, sla, pri, "", "", ""]
        for col, value in enumerate(values, start=2):
            ws.cell(r, col, value)
            ws.cell(r, col).border = THIN_BORDER
        r += 1

    r += 1
    ws.merge_cells(f"B{r}:H{r}")
    ws[f"B{r}"] = "B. ALLOCATION PRIORITY (WHEN POOL IS CONSTRAINED)"
    _style_range(ws, r, 2, 8, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    r += 1
    priority_data = [
        ("1st (Highest)", "MT large accounts (supermarket chains, confirmed contractual POs)"),
        ("2nd", "MT Ecommerce (2-4h SLA, platform rating at risk)"),
        ("3rd", "TT large distributors (high-volume, long-term relationship)"),
        ("4th (Lowest)", "TT small / walk-in customers (flexible SLA, served last)"),
    ]
    for pri, desc in priority_data:
        ws.cell(r, 2, pri)
        ws.cell(r, 2).font = BOLD_FONT
        ws.cell(r, 3, desc)
        _style_range(ws, r, 2, 8, alignment=LEFT)
        r += 1

    r += 1
    ws.merge_cells(f"B{r}:H{r}")
    ws[f"B{r}"] = "C. OPERATIONAL MECHANISMS"
    _style_range(ws, r, 2, 8, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    r += 1
    ops_data = [
        ("R1 Common Pool", "All SKU stock merged into single WMS pool; OMS tracks virtual availability per channel"),
        ("R2 Per-Channel ROP Floor", "Each channel has its own ROP threshold; replenishment triggers when any channel hits its floor"),
        ("R3 Amber Alert", "Pool < sum of all channel ROPs raise alert, prioritize confirmed MT POs"),
        ("R4 Red Conflict", "Pool < total SS activate priority allocation order (MT first, TT last) + emergency PO"),
        ("R5 AZ Spike Freeze", "Demand-spike AZ SKUs: temporarily reserve stock for MT, freeze TT, trigger emergency PO"),
        ("R6 End-of-Day Rebalance", "Reset channel virtual allocations to ROP-proportional targets using rolling 7-day demand forecast"),
    ]
    for rule, desc in ops_data:
        ws.cell(r, 2, rule)
        ws.cell(r, 2).font = BOLD_FONT
        ws.cell(r, 3, desc)
        ws.merge_cells(f"C{r}:H{r}")
        _style_range(ws, r, 2, 8, alignment=LEFT)
        r += 1

    r += 1
    ws.merge_cells(f"B{r}:H{r}")
    ws[f"B{r}"] = "D. QUANTIFICATION POOLING IMPACT ON SAFETY STOCK"
    _style_range(ws, r, 2, 8, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    r += 1
    headers = ["Scenario", "Formula", "Total SS", "Savings %", "Reduction Factor", "", "", ""]
    for col, value in enumerate(headers, start=2):
        ws.cell(r, col, value)
    _style_range(ws, r, 2, 8, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    r += 1
    for _, row in inventory_pooling_summary.iterrows():
        values = [
            row.get("scenario", ""),
            row.get("formula", ""),
            round(row["total_ss"], 0) if pd.notna(row.get("total_ss")) else 0,
            row["savings_pct"] if pd.notna(row.get("savings_pct")) else 0,
            f"{(1 - row['savings_pct']):.2f}" if pd.notna(row.get("savings_pct")) and row["savings_pct"] < 1 else "N/A",
        ]
        for col, value in enumerate(values, start=2):
            cell = ws.cell(r, col, value)
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            if col == 4:
                cell.number_format = "#,##0"
            elif col == 5:
                cell.number_format = "0.0%"
        r += 1

    pool_row = inventory_pooling_summary[
        inventory_pooling_summary["scenario"] == "Pooled (Mile-Weighted)"
    ]
    if not pool_row.empty:
        r += 1
        savings_val = pool_row.iloc[0]["savings_pct"]
        ws.cell(r, 2, "KEY RESULT:")
        ws.cell(r, 2).font = BOLD_FONT
        ws.cell(r, 3, f"Virtual pooling with mile-weighted LT saves approximately {savings_val:.1%} safety stock vs separated channels")
        ws.merge_cells(f"C{r}:H{r}")
        _style_range(ws, r, 2, 8, fill=TITLE_FILL, font=Font(bold=True, color="FFFFFF"), alignment=LEFT)

    r += 3
    ws.merge_cells(f"B{r}:L{r}")
    ws[f"B{r}"] = "E. SAFETY STOCK PER CLASS A SKU"
    _style_range(ws, r, 2, 12, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    
    r += 1
    ws.merge_cells(f"B{r}:L{r}")
    ws[f"B{r}"] = "FORMULAS:  SS = Z x sigma_daily x sqrt(LT_avg)    |    ROP = mu_daily x LT_avg + SS"
    _style_range(ws, r, 2, 12, fill=SUBHEADER_FILL, font=BOLD_FONT, alignment=LEFT)
    
    r += 1
    ws.merge_cells(f"B{r}:L{r}")
    ws[f"B{r}"] = "Parameters: Z = 1.645 (95% service level), LT_avg = 1.94 days, sqrt(LT) = 1.393, ddof=1 over 7 months (Jun Dec 2025)"
    _style_range(ws, r, 2, 12, fill=SUBHEADER_FILL, alignment=LEFT)
    
    r += 1
    headers_ss = [
        "SKU", "Product", "ABC-XYZ", "mu_monthly", "sigma_monthly",
        "mu_daily", "sigma_daily", "Z", "sqrt(LT)", "Safety Stock", "ROP", ""
    ]
    for col, value in enumerate(headers_ss, start=2):
        ws.cell(r, col, value)
    _style_range(ws, r, 2, 12, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    r += 1

    for _, row_data in safety_stock_class_a.iterrows():
        abc_xyz_code = f"{row_data.get('abc_quantity', '')}{row_data.get('abc_frequency', '')}"
        values = [
            row_data["sku_code"],
            row_data.get("product_name", ""),
            abc_xyz_code,
            row_data["mu_monthly"] if pd.notna(row_data["mu_monthly"]) else 0,
            row_data["sigma_monthly"] if pd.notna(row_data["sigma_monthly"]) else 0,
            row_data["mu_daily"] if pd.notna(row_data["mu_daily"]) else 0,
            row_data["sigma_daily"] if pd.notna(row_data["sigma_daily"]) else 0,
            1.645,
            row_data["sqrt_lt"] if pd.notna(row_data["sqrt_lt"]) else 0,
            row_data["ss"] if pd.notna(row_data["ss"]) else 0,
            row_data["rop"] if pd.notna(row_data["rop"]) else 0,
        ]
        for col, value in enumerate(values, start=2):
            cell = ws.cell(r, col, value)
            cell.border = THIN_BORDER
            cell.alignment = LEFT if col in {2, 3, 4} else CENTER
            if col in {5, 6, 7, 8, 10, 11, 12} and isinstance(value, (int, float)):
                cell.number_format = "0.0"
            elif col == 11:
                cell.number_format = "0"
        r += 1

    ws.freeze_panes = "B3"


def _write_q31_slotting_summary_sheet(ws, travel_metrics: dict) -> None:
    widths = [3, 30, 20, 14, 16, 20, 20, 16, 16]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    _apply_title(ws, "B1", 8, "LOGage 2026 — QUESTION 3.1 | SLOTTING OPTIMIZATION & TRAVEL PROOF")

    r = 3
    ws.merge_cells(f"B{r}:H{r}")
    ws[f"B{r}"] = "A. MODEL ASSUMPTIONS"
    _style_range(ws, r, 2, 8, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    r += 1
    headers = ["Parameter", "Value", "Unit", "Notes", "", "", "", ""]
    for col, value in enumerate(headers, start=2):
        ws.cell(r, col, value)
    _style_range(ws, r, 2, 8, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    r += 1
    assumptions = [
        ("Walk Speed", travel_metrics.get("walk_speed_m_min", 72), "m/min", "1.2 m/s standard"),
        ("Round-Trip Factor", travel_metrics.get("round_trip_factor", 2), "", "Pick location to packing and back"),
        ("Zone A Distance to Packing", travel_metrics.get("dist_a_m", 8), "m", "Golden zone, nearest"),
        ("Zone B Distance to Packing", travel_metrics.get("dist_b_m", 25), "m", "Mid bay"),
        ("Zone C Distance to Packing", travel_metrics.get("dist_c_m", 47.5), "m", "Back / upper rack"),
        ("Pick Rate A (Pick-Face)", 4, "pcs/min", "Fast pick-face"),
        ("Pick Rate B (Forward Reserve)", 2.5, "pcs/min", "Carton-centric"),
        ("Pick Rate C (Reserve/Bulk)", 2, "pcs/min", "Slow reserve"),
    ]
    for param, val, unit, note in assumptions:
        values = [param, val, unit, note, "", "", "", ""]
        for col, value in enumerate(values, start=2):
            cell = ws.cell(r, col, value)
            cell.border = THIN_BORDER
            cell.alignment = LEFT if col in {2, 5} else CENTER
        r += 1

    r += 2
    ws.merge_cells(f"B{r}:H{r}")
    ws[f"B{r}"] = "B. ZONE DESIGN"
    _style_range(ws, r, 2, 8, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    r += 1
    headers_zone = ["Zone", "Class", "# SKUs", "Total Picks", "Pick Mode", "Slot Rule", "Dist to Packing (m)", ""]
    for col, value in enumerate(headers_zone, start=2):
        ws.cell(r, col, value)
    _style_range(ws, r, 2, 8, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    r += 1
    zone_data = [
        ("A Pick-Face (Golden)", "A / Fast", travel_metrics["zone_counts"]["A"], travel_metrics["zone_picks"]["A"],
         "Loose + carton break-bulk", "Highest order freq nearest slots", travel_metrics.get("dist_a_m", 8)),
        ("B Forward Reserve", "B / Medium", travel_metrics["zone_counts"]["B"], travel_metrics["zone_picks"]["B"],
         "Carton picking + replenish A", "Medium velocity mid bay", travel_metrics.get("dist_b_m", 25)),
        ("C Reserve / Bulk", "C / Slow", travel_metrics["zone_counts"]["C"], travel_metrics["zone_picks"]["C"],
         "Pallet store, pick on demand", "Low velocity far/upper rack", travel_metrics.get("dist_c_m", 47.5)),
    ]
    for name, cls, count, picks, mode, rule, dist in zone_data:
        values = [name, cls, count, picks, mode, rule, dist]
        for col, value in enumerate(values, start=2):
            cell = ws.cell(r, col, value)
            cell.border = THIN_BORDER
            cell.alignment = LEFT if col in {2, 6, 7} else CENTER
        r += 1

    r += 2
    ws.merge_cells(f"B{r}:H{r}")
    ws[f"B{r}"] = "C. ZONE DISTANCE TABLE"
    _style_range(ws, r, 2, 8, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    r += 1
    headers_dist = ["Zone", "Picks (T_ij)", "Dist D_ij (m)", "Round-Trip Dist (m)", "Pick Rate (pcs/min)", "Travel Time (min)", "", ""]
    for col, value in enumerate(headers_dist, start=2):
        ws.cell(r, col, value)
    _style_range(ws, r, 2, 8, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    r += 1
    walk_speed = travel_metrics.get("walk_speed_m_min", 72)
    for zone_detail in travel_metrics.get("zone_detail", []):
        z = zone_detail["zone"]
        picks = zone_detail["picks"]
        dist = zone_detail["dist_m"]
        rt_dist = zone_detail["round_trip_dist_m"]
        pick_rate = {"A": 4, "B": 2.5, "C": 2}.get(z, 2)
        travel_time = rt_dist / walk_speed if walk_speed else 0
        values = [f"Zone {z}", picks, dist, round(rt_dist, 0), pick_rate, round(travel_time, 1)]
        for col, value in enumerate(values, start=2):
            cell = ws.cell(r, col, value)
            cell.border = THIN_BORDER
            cell.alignment = LEFT if col == 2 else CENTER
            if col in {4, 5, 7}:
                cell.number_format = "#,##0"
        r += 1

    total_picks = travel_metrics.get("total_picks", 0)
    total_rt = travel_metrics.get("opt_total_round_trip_m", 0)
    total_time = travel_metrics.get("opt_travel_time_min", 0)
    values_tot = ["TOTAL", total_picks, "", round(total_rt, 0), "", round(total_time, 1)]
    for col, value in enumerate(values_tot, start=2):
        cell = ws.cell(r, col, value)
        cell.border = THIN_BORDER
        cell.font = BOLD_FONT
        cell.alignment = CENTER
    _style_range(ws, r, 2, 7, fill=SUBHEADER_FILL, font=BOLD_FONT, alignment=CENTER)
    r += 2

    ws.merge_cells(f"B{r}:H{r}")
    ws[f"B{r}"] = "D. RESULTS OPTIMIZED vs BASELINE"
    _style_range(ws, r, 2, 8, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    r += 1
    headers_res = ["Metric", "Baseline (Random)", "Optimized (ABC-Zoned)", "Improvement", "Status", "", "", ""]
    for col, value in enumerate(headers_res, start=2):
        ws.cell(r, col, value)
    _style_range(ws, r, 2, 8, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    r += 1
    meets_target = travel_metrics.get("meets_target", False)
    reduction = travel_metrics.get("travel_reduction", 0)
    target = travel_metrics.get("reduction_target", 0.3)
    results = [
        ("Avg One-Way Distance/Pick (m)", travel_metrics.get("baseline_avg_dist_m", 0), travel_metrics.get("opt_avg_oneway_m_per_pick", 0),
         f"{reduction * 100:.1f}% reduction", "MEETS TARGET" if meets_target else "BELOW TARGET"),
        ("Total Round-Trip Distance (m)", travel_metrics.get("baseline_total_round_trip_m", 0), travel_metrics.get("opt_total_round_trip_m", 0),
         f"{reduction * 100:.1f}% reduction", ""),
        ("Travel Time (min)", travel_metrics.get("baseline_travel_time_min", 0), travel_metrics.get("opt_travel_time_min", 0),
         f"{reduction * 100:.1f}% reduction", ""),
        ("Travel Time Reduction %", "", f"{reduction * 100:.1f}%", f"Target >= {target * 100:.0f}%", ""),
    ]
    for metric, base, opt, improvement, status in results:
        values = [metric, base, opt, improvement, status, "", ""]
        for col, value in enumerate(values, start=2):
            cell = ws.cell(r, col, value)
            cell.border = THIN_BORDER
            cell.alignment = LEFT if col in {2, 6} else CENTER
            if col in {3, 4} and isinstance(value, float):
                cell.number_format = "#,##0.0"
        if metric == "Travel Time Reduction %":
            fill_color = "D9EAD3" if meets_target else "F4CCCC"
            _style_range(ws, r, 2, 7, fill=PatternFill("solid", fgColor=fill_color), font=BOLD_FONT, alignment=None)
        r += 1
    ws.freeze_panes = "B3"


def _write_q31_slot_assignment_sheet(ws, slotting_plan: pd.DataFrame) -> None:
    for idx, width in enumerate([3, 10, 14, 22, 12, 10, 14, 14, 16, 14, 14, 12], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    _apply_title(ws, "B1", 12, "LOGage 2026 — QUESTION 3.1 | SLOT ASSIGNMENT")

    headers = [
        "Slot ID", "SKU Code", "Product", "Cat", "Zone",
        "Order Freq", "Pick Rate (pcs/min)", "Velocity (Freq/Rate)",
        "Size (L/pcs)", "Avg Order Qty", "Dist (m)", ""
    ]
    for col, value in enumerate(headers, start=2):
        ws.cell(2, col, value)
    _style_range(ws, 2, 2, 12, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)

    zone_order = {"Pick-Face Zone (Ground Level)": "A", "Forward Reserve Zone": "B", "Reserve / Bulk Zone": "C"}
    r = 3
    for _, row in slotting_plan.iterrows():
        zone_code = zone_order.get(row.get("zone_assignment", ""), "")
        slot_id = f"{zone_code}-{row.get('rank_in_zone', 0):03d}"
        pick_rate = row.get("pick_rate_pcs_min", 0)
        velocity = row.get("velocity", 0)
        cbm = row.get("cbm_incl_flap_m3", 0)
        size_l = cbm * 1000 if pd.notna(cbm) else 0
        values = [
            slot_id,
            row.get("sku_code", ""),
            row.get("product_name", "") if "product_name" in row else row.get("sku_code", ""),
            row.get("abc_class", ""),
            zone_code,
            int(row.get("travel_contribution", 0)) if pd.notna(row.get("travel_contribution")) else 0,
            round(pick_rate, 1),
            round(velocity, 1),
            round(size_l, 2),
            round(row.get("avg_qty_per_order", 0), 1) if pd.notna(row.get("avg_qty_per_order")) else 0,
            row.get("slot_distance_m", 0),
        ]
        for col, value in enumerate(values, start=2):
            cell = ws.cell(r, col, value)
            cell.border = THIN_BORDER
            cell.alignment = LEFT if col in {3, 4, 5} else CENTER
            if col in {8, 9} and isinstance(value, float):
                cell.number_format = "0.0"
            elif col == 10:
                cell.number_format = "0.00"
        cls = row.get("abc_class", "")
        if cls in CLASS_FILLS:
            _style_range(ws, r, 2, 12, fill=CLASS_FILLS[cls], alignment=None)
        r += 1
    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"B2:{get_column_letter(12)}{r - 1}"


def _write_q31_u_shape_sheet(ws) -> None:
    for idx, width in enumerate([3, 18, 18, 18, 18, 18, 18, 18], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    _apply_title(ws, "B1", 8, "LOGage 2026 — QUESTION 3.1 | U-SHAPE WAREHOUSE HEATMAP")

    r = 3
    ws.merge_cells(f"B{r}:H{r}")
    ws[f"B{r}"] = (
        "U-Shaped (Horse-Shoe) Layout: Receiving & Packing share the same dock wall. "
        "Zone A (Golden Pick-Face) is the nearest zone to the packing station, "
        "followed by Zone B (Forward Reserve), then Zone C (Reserve/Bulk) at the far end."
    )
    ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="center")
    _style_range(ws, r, 2, 8, fill=SUBHEADER_FILL, alignment=LEFT)
    r += 2

    ws.merge_cells(f"B{r}:H{r}")
    ws[f"B{r}"] = "Layout Diagram:"
    ws[f"B{r}"].font = BOLD_FONT
    r += 1
    layout_lines = [
        "+---------------------------------------------------+",
        "|                  DOCK (RECEIVING / SHIPPING)       |",
        "+---------+---------+---------+---------+----------+",
        "| ZONE A  | ZONE A  | PACKING | ZONE A  | ZONE A   |",
        "| Pick-Fc | Pick-Fc | STATION | Pick-Fc | Pick-Fc  |",
        "| (Ground)| (Ground)|         | (Ground)| (Ground) |",
        "+---------+---------+---------+---------+----------+",
        "| ZONE B (Forward Reserve Mid Bay)                  |",
        "+---------------------------------------------------+",
        "| ZONE C (Reserve / Bulk Back Wall & Upper Racks)    |",
        "+---------------------------------------------------+",
    ]
    for line in layout_lines:
        ws.cell(r, 2, line)
        ws.cell(r, 2).font = Font(name="Courier New", size=9)
        r += 1

    img_path = CHARTS_DIR / "q31_u_shape_heatmap.png"
    if img_path.exists():
        try:
            img = XLImage(str(img_path))
            orig_w, orig_h = img.width, img.height
            img.width = 650
            img.height = int(650 * orig_h / orig_w)
            ws.add_image(img, f"B{r + 4}")
        except Exception as e:
            print(f"Warning: could not embed Q3.1 heatmap image: {e}")

    ws.freeze_panes = "B3"


def _write_q32_pick_pack_sheet(ws) -> None:
    for idx, width in enumerate([3, 18, 30, 30, 28, 30], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    _apply_title(ws, "B1", 6, "LOGage 2026 — QUESTION 3.2 | PICK & PACK PROCESS")

    r = 3
    ws.merge_cells(f"B{r}:F{r}")
    ws[f"B{r}"] = "A. TWO PICK MODELS"
    _style_range(ws, r, 2, 6, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    r += 1
    headers = ["Channel", "Order Profile", "Pick Model", "Mechanics", "Pack"]
    for col, value in enumerate(headers, start=2):
        ws.cell(r, col, value)
    _style_range(ws, r, 2, 6, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    r += 1
    models = [
        (
            "B2B",
            "Large volume, few SKUs",
            "Pallet / Total Picking",
            "One picker fulfils whole pallets or large CTN qty per order in a single pass. Sources Zone A pick-face + Zone C bulk.",
            "Stage whole pallets; minimal re-sort.",
        ),
        (
            "B2C / E-com",
            "Small qty, many SKUs, many concurrent orders",
            "Batch Picking + Zone Routing",
            "Group many small orders into one pick wave; pickers route by zone; items consolidated at sortation.",
            "Sort wave back to individual orders at packing.",
        ),
    ]
    for ch, profile, model, mechanics, pack in models:
        values = [ch, profile, model, mechanics, pack]
        for col, value in enumerate(values, start=2):
            cell = ws.cell(r, col, value)
            cell.border = THIN_BORDER
            cell.alignment = LEFT
            cell.font = Font(size=10)
        r += 1

    r += 2
    ws.merge_cells(f"B{r}:F{r}")
    ws[f"B{r}"] = "B. SHARED-SKU CONFLICT ALLOCATION RULES"
    _style_range(ws, r, 2, 6, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    r += 1
    headers = ["Rule", "Condition", "Action", "", "", ""]
    for col, value in enumerate(headers, start=2):
        ws.cell(r, col, value)
    _style_range(ws, r, 2, 6, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    r += 1
    rules = [
        ("Check", "Same SKU, both channels, concurrent", "Evaluate on-hand vs (B2B + B2C) demand"),
        ("Sufficient", "On-hand >= B2B + B2C", "Allocate to both from shared pool; no rationing"),
        ("R1", "Short stock", "Commit to the confirmed B2B PO first (contractual, penalty risk)"),
        ("R2", "Short stock", "Hold the B2C safety buffer so the storefront does not show stockout"),
        ("R3", "Short + B2C SLA-critical (2-4h)", "Split: B2C SLA order gets partial now, backorder remainder"),
        ("R4", "Below reorder point", "Trigger emergency replenishment / inter-warehouse transfer"),
    ]
    for rule, condition, action in rules:
        values = [rule, condition, action, "", "", ""]
        for col, value in enumerate(values, start=2):
            cell = ws.cell(r, col, value)
            cell.border = THIN_BORDER
            cell.alignment = LEFT
            cell.font = Font(size=10)
        r += 1

    img_path = CHARTS_DIR / "q32_pick_pack_flowchart.png"
    if img_path.exists():
        try:
            img = XLImage(str(img_path))
            orig_w, orig_h = img.width, img.height
            img.width = 550
            img.height = int(550 * orig_h / orig_w)
            ws.add_image(img, f"B{r + 5}")
        except Exception as e:
            print(f"Warning: could not embed Q3.2 flowchart image: {e}")

    ws.freeze_panes = "B3"
