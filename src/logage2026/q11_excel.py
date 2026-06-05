from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.logage2026.config import Q11_WORKBOOK_OUTPUT


THIN_BORDER = Border(
    left=Side(style="thin", color="D9DEE5"),
    right=Side(style="thin", color="D9DEE5"),
    top=Side(style="thin", color="D9DEE5"),
    bottom=Side(style="thin", color="D9DEE5"),
)
TITLE_FILL = PatternFill("solid", fgColor="0F4C81")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="EAF2F8")
CLASS_FILLS = {
    "A": PatternFill("solid", fgColor="DDEBF7"),
    "B": PatternFill("solid", fgColor="FFF2CC"),
    "C": PatternFill("solid", fgColor="F4CCCC"),
    "X": PatternFill("solid", fgColor="D9EAD3"),
    "Y": PatternFill("solid", fgColor="FCE5CD"),
    "Z": PatternFill("solid", fgColor="F4CCCC"),
}
TITLE_FONT = Font(bold=True, color="FFFFFF", size=14)
SECTION_FONT = Font(bold=True, color="1F1F1F", size=11)
HEADER_FONT = Font(bold=True, color="FFFFFF")
BOLD_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def write_q11_workbook(
    abc_xyz: pd.DataFrame,
    abc_xyz_matrix: pd.DataFrame,
    monthly_demand: pd.DataFrame,
    q11_shipments: pd.DataFrame,
    output_path: Path = Q11_WORKBOOK_OUTPUT,
) -> Path:
    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "Dashboard"
    _write_dashboard(dashboard, abc_xyz, abc_xyz_matrix, q11_shipments)
    _write_full_sku_ranking(workbook.create_sheet("Full SKU Ranking"), abc_xyz)
    _write_monthly_demand_sheet(workbook.create_sheet("Monthly Demand (XYZ Base)"), monthly_demand)
    _write_class_a_deep_dive(workbook.create_sheet("Class A Deep Dive"), monthly_demand)
    workbook.save(output_path)
    return output_path


def _apply_title(ws, start_cell: str, end_col: int, text: str) -> None:
    row = ws[start_cell].row
    start_col = ws[start_cell].column
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws[start_cell]
    cell.value = text
    cell.fill = TITLE_FILL
    cell.font = TITLE_FONT
    cell.alignment = LEFT


def _style_range(ws, row: int, start_col: int, end_col: int, *, fill=None, font=None, alignment=None) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.border = THIN_BORDER
        if fill is not None:
            cell.fill = fill
        if font is not None:
            cell.font = copy(font)
        if alignment is not None:
            cell.alignment = alignment


def _format_share(value: float) -> str:
    return f"{value:.1%}"


def _safe_float(value: object) -> float:
    return 0.0 if pd.isna(value) else float(value)


def _period_label(start: pd.Timestamp, end: pd.Timestamp) -> str:
    if start.year == end.year:
        return f"{start.strftime('%b')}–{end.strftime('%b %Y')}"
    return f"{start.strftime('%b %Y')}–{end.strftime('%b %Y')}"


def _month_columns(monthly_demand: pd.DataFrame) -> list[str]:
    month_like = []
    for column in monthly_demand.columns:
        if column in {"sku_code", "product_name", "category", "abc_quantity", "abc_frequency", "xyz", "abc_xyz", "Total", "Mean/Mo", "Std Dev", "CV", "quantity", "quantity_share", "order_frequency", "frequency_share", "cbm_total"}:
            continue
        try:
            pd.Period(column, freq="M")
            month_like.append(column)
        except Exception:
            continue
    return month_like


def _operational_category(abc_xyz_code: str) -> str:
    mapping = {
        "AX": "Fast & Predictable",
        "AY": "Fast & Predictable",
        "AZ": "Fast & Volatile",
        "BX": "Medium & Stable",
        "BY": "Medium & Seasonal",
        "BZ": "Medium & Volatile",
        "CX": "Slow & Stable",
        "CY": "Slow & Seasonal",
        "CZ": "Slow & Erratic",
    }
    return mapping.get(abc_xyz_code, "Unclassified")


def _slotting_zone(abc_class: str) -> str:
    if abc_class == "A":
        return "Pick-Face Zone (Ground Level)"
    if abc_class == "B":
        return "Forward Reserve Zone"
    return "Reserve / Bulk Zone"


def _demand_pattern(xyz_class: str) -> str:
    return {"X": "Stable", "Y": "Seasonal", "Z": "Erratic"}.get(xyz_class, "Unknown")


def _stock_strategy(xyz_class: str) -> str:
    return {
        "X": "Fixed ROP + EOQ",
        "Y": "Dynamic MRP + safety stock",
        "Z": "Buffer stock + safety stock",
    }.get(xyz_class, "Manual review")


def _reorder_logic(xyz_class: str) -> str:
    return {"X": "Min-Max fixed", "Y": "Weekly forecast", "Z": "Safety stock + review"}.get(xyz_class, "Manual review")


def _dashboard_abc_summary(abc_xyz: pd.DataFrame) -> pd.DataFrame:
    summary = (
        abc_xyz.groupby("abc_quantity", dropna=False)
        .agg(sku_count=("sku_code", "nunique"), total_qty=("quantity", "sum"), total_cbm=("cbm_total", "sum"))
        .reset_index()
    )
    summary["sku_share"] = summary["sku_count"] / abc_xyz["sku_code"].nunique()
    summary["qty_share"] = summary["total_qty"] / abc_xyz["quantity"].sum()
    return summary.set_index("abc_quantity").reindex(["A", "B", "C"]).reset_index()


def _dashboard_xyz_summary(abc_xyz: pd.DataFrame) -> pd.DataFrame:
    summary = (
        abc_xyz.groupby("xyz", dropna=False)
        .agg(sku_count=("sku_code", "nunique"), avg_cv=("demand_cv", "mean"))
        .reset_index()
    )
    summary["sku_share"] = summary["sku_count"] / abc_xyz["sku_code"].nunique()
    return summary.set_index("xyz").reindex(["X", "Y", "Z"]).reset_index()


def _write_dashboard(
    ws,
    abc_xyz: pd.DataFrame,
    abc_xyz_matrix: pd.DataFrame,
    q11_shipments: pd.DataFrame,
) -> None:
    period_start = q11_shipments["created_date"].min()
    period_end = q11_shipments["created_date"].max()
    month_count = q11_shipments["created_date"].dt.to_period("M").nunique()
    for idx, width in enumerate([3, 18, 14, 12, 16, 14, 12, 24, 40], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    _apply_title(ws, "B1", 9, "LOGage 2026 — QUESTION 1.1 | ABC-XYZ ANALYSIS DASHBOARD")
    period_text = (
        f"Data: My Phuoc & Vinh Loc Warehouses  |  Period: {_period_label(period_start, period_end)}  |  "
        f"{len(q11_shipments):,} transaction lines  |  {abc_xyz['sku_code'].nunique():,} unique SKUs"
    )
    ws.merge_cells("B2:I2")
    ws["B2"] = period_text
    ws["B2"].alignment = LEFT
    ws["B2"].font = Font(italic=True, color="5B5B5B")

    abc_summary = _dashboard_abc_summary(abc_xyz)
    ws.merge_cells("B4:I4")
    ws["B4"] = "A.  ABC CLASSIFICATION — OUTBOUND QUANTITY (threshold: A=70%, B=90%, C=100%)"
    _style_range(ws, 4, 2, 9, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    headers = ["Class", "SKU Count", "% of SKUs", "Total Qty (pcs)", "% of Total Qty", "Total CBM", "Threshold", "Action"]
    for col, value in enumerate(headers, start=2):
        ws.cell(5, col, value)
    _style_range(ws, 5, 2, 9, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    thresholds = {
        "A": "Cum. top 70% volume",
        "B": "70-90% cum. volume",
        "C": "Bottom 10% volume",
    }
    actions = {
        "A": "Pick-face zone | Daily cycle count | Min lead time",
        "B": "Intermediate zone | Weekly review",
        "C": "Reserve/bulk zone | Monthly review",
    }
    for row_idx, row in enumerate(abc_summary.itertuples(index=False), start=6):
        values = [
            row.abc_quantity,
            int(row.sku_count),
            _format_share(row.sku_share),
            round(_safe_float(row.total_qty)),
            _format_share(row.qty_share),
            round(_safe_float(row.total_cbm)),
            thresholds[row.abc_quantity],
            actions[row.abc_quantity],
        ]
        for col, value in enumerate(values, start=2):
            ws.cell(row_idx, col, value)
        _style_range(ws, row_idx, 2, 9, fill=CLASS_FILLS[row.abc_quantity], alignment=LEFT)

    xyz_summary = _dashboard_xyz_summary(abc_xyz)
    ws.merge_cells("B10:I10")
    ws["B10"] = f"B.  XYZ CLASSIFICATION — DEMAND VARIABILITY (Coefficient of Variation over {month_count} months)"
    _style_range(ws, 10, 2, 9, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    xyz_headers = ["Class", "SKU Count", "% of SKUs", "CV Range", "Avg CV", "Demand Pattern", "Replenishment Logic", "Safety Stock"]
    for col, value in enumerate(xyz_headers, start=2):
        ws.cell(11, col, value)
    _style_range(ws, 11, 2, 9, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    cv_ranges = {"X": "CV <= 0.50", "Y": "0.50 < CV <= 1.00", "Z": "CV > 1.00"}
    repl = {"X": "Fixed reorder point (ROP)", "Y": "Dynamic MRP/forecast-based", "Z": "Demand-driven + buffer stock"}
    safety = {"X": "Low — 1-2 weeks cover", "Y": "Medium — 2-4 weeks cover", "Z": "High — 4-8 weeks cover"}
    demand = {"X": "Stable / consistent", "Y": "Seasonal or trend", "Z": "Erratic / unpredictable"}
    for row_idx, row in enumerate(xyz_summary.itertuples(index=False), start=12):
        values = [
            row.xyz,
            int(row.sku_count),
            _format_share(row.sku_share),
            cv_ranges[row.xyz],
            round(_safe_float(row.avg_cv), 2),
            demand[row.xyz],
            repl[row.xyz],
            safety[row.xyz],
        ]
        for col, value in enumerate(values, start=2):
            ws.cell(row_idx, col, value)
        _style_range(ws, row_idx, 2, 9, fill=CLASS_FILLS[row.xyz], alignment=LEFT)

    ws.merge_cells("B17:I17")
    ws["B17"] = "C.  COMBINED ABC-XYZ MATRIX — SKU Count"
    _style_range(ws, 17, 2, 9, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    matrix = (
        abc_xyz_matrix.pivot(index="abc_quantity", columns="xyz", values="sku_count")
        .reindex(index=["A", "B", "C"], columns=["X", "Y", "Z"])
        .fillna(0)
        .astype(int)
    )
    ws.append([])
    headers = ["", "X (Stable)", "Y (Seasonal)", "Z (Erratic)", "Row Total", "% Volume", "% Freq", "Key Implication"]
    for col, value in enumerate(headers, start=2):
        ws.cell(18, col, value)
    _style_range(ws, 18, 2, 9, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    volume_share = abc_xyz.groupby("abc_quantity")["quantity"].sum() / abc_xyz["quantity"].sum()
    freq_share = abc_xyz.groupby("abc_quantity")["order_frequency"].sum() / abc_xyz["order_frequency"].sum()
    implications = {
        "A": "HIGH PRESSURE — Daily ops, premium slots, SLA-critical",
        "B": "MODERATE — Standard fulfillment, replenishment planning",
        "C": "LOW PRIORITY — Bulk storage, periodic review",
    }
    for row_idx, abc_class in enumerate(["A", "B", "C"], start=19):
        values = [
            abc_class,
            int(matrix.loc[abc_class, "X"]),
            int(matrix.loc[abc_class, "Y"]),
            int(matrix.loc[abc_class, "Z"]),
            int(matrix.loc[abc_class].sum()),
            _format_share(_safe_float(volume_share.get(abc_class))),
            _format_share(_safe_float(freq_share.get(abc_class))),
            implications[abc_class],
        ]
        for col, value in enumerate(values, start=2):
            ws.cell(row_idx, col, value)
        _style_range(ws, row_idx, 2, 9, fill=CLASS_FILLS[abc_class], alignment=LEFT)
    total_row = 22
    total_values = ["TOTAL", int(matrix["X"].sum()), int(matrix["Y"].sum()), int(matrix["Z"].sum()), int(matrix.values.sum()), "", "", ""]
    for col, value in enumerate(total_values, start=2):
        ws.cell(total_row, col, value)
    _style_range(ws, total_row, 2, 9, fill=SUBHEADER_FILL, font=BOLD_FONT, alignment=CENTER)

    class_a = abc_xyz[abc_xyz["abc_quantity"].eq("A")].copy()
    ws.merge_cells("B24:I24")
    ws["B24"] = "D.  FAST-MOVING SKU GROUP (Class A) — Primary Source of Warehouse Operational Pressure"
    _style_range(ws, 24, 2, 9, fill=SECTION_FILL, font=SECTION_FONT, alignment=LEFT)
    headers = ["Metric", "AX — Stable Fast", "AY — Seasonal Fast", "AZ — Volatile Fast", "ALL Class A Total", "% of Grand Total", "Operational Priority"]
    for col, value in enumerate(headers, start=2):
        ws.cell(25, col, value)
    _style_range(ws, 25, 2, 8, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    subgroup_values = {}
    for subgroup in ["AX", "AY", "AZ"]:
        subset = class_a[class_a["abc_xyz"].eq(subgroup)]
        subgroup_values[subgroup] = {
            "sku_count": int(subset["sku_code"].nunique()),
            "quantity": subset["quantity"].sum(),
            "order_frequency": subset["order_frequency"].sum(),
            "avg_cv": subset["demand_cv"].mean(),
            "cbm_total": subset["cbm_total"].sum(),
        }
    metrics = [
        ("SKU Count", "sku_count", f"{_format_share(class_a['sku_code'].nunique() / abc_xyz['sku_code'].nunique())} of all SKUs", "Fast-moving concentration remains high"),
        ("Total Qty (pcs)", "quantity", f"{_format_share(class_a['quantity'].sum() / abc_xyz['quantity'].sum())} of total volume", "Prioritize in pick-face zone"),
        ("Order Frequency", "order_frequency", f"{_format_share(class_a['order_frequency'].sum() / abc_xyz['order_frequency'].sum())} of orders", "High pick frequency -> batch picking"),
        ("Avg Monthly Demand", None, "-", f"{month_count}-month average"),
        ("Avg CV (Variability)", "avg_cv", "-", "High safety stock needed"),
        ("Total CBM", "cbm_total", f"{_format_share(class_a['cbm_total'].sum() / abc_xyz['cbm_total'].sum())} of total CBM", "Slotting space priority"),
    ]
    for row_idx, (metric, key, pct_text, priority_text) in enumerate(metrics, start=26):
        ws.cell(row_idx, 2, metric)
        for col_idx, subgroup in enumerate(["AX", "AY", "AZ"], start=3):
            value = subgroup_values[subgroup].get(key) if key else "-"
            if key in {"quantity", "order_frequency", "cbm_total"}:
                value = f"{round(_safe_float(value)):,}"
            elif key == "avg_cv":
                value = f"{_safe_float(value):.2f}"
            elif key == "sku_count":
                value = str(value)
            ws.cell(row_idx, col_idx, value)
        total_value = "-"
        if key == "sku_count":
            total_value = str(int(class_a["sku_code"].nunique()))
        elif key == "quantity":
            total_value = f"{round(class_a['quantity'].sum()):,}"
        elif key == "order_frequency":
            total_value = f"{round(class_a['order_frequency'].sum()):,}"
        elif key == "avg_cv":
            total_value = f"{class_a['demand_cv'].mean():.2f}"
        elif key == "cbm_total":
            total_value = f"{round(class_a['cbm_total'].sum()):,}"
        elif metric == "Avg Monthly Demand":
            total_value = f"{round(class_a['quantity'].sum() / month_count):,}"
        ws.cell(row_idx, 6, total_value)
        ws.cell(row_idx, 7, pct_text)
        ws.cell(row_idx, 8, priority_text)
        _style_range(ws, row_idx, 2, 8, fill=SUBHEADER_FILL if row_idx % 2 == 0 else None, alignment=LEFT)

    predictable = int(class_a["abc_xyz"].isin(["AX", "AY"]).sum())
    volatile = int(class_a["abc_xyz"].eq("AZ").sum())
    key_finding = (
        f"KEY FINDING: {int(class_a['sku_code'].nunique())} Class A SKUs "
        f"({_format_share(class_a['sku_code'].nunique() / abc_xyz['sku_code'].nunique())} of portfolio) account for "
        f"{_format_share(class_a['quantity'].sum() / abc_xyz['quantity'].sum())} of outbound volume and "
        f"{_format_share(class_a['order_frequency'].sum() / abc_xyz['order_frequency'].sum())} of order frequency. "
        f"Of these, {predictable} SKUs (AX+AY) are predictable enough for fixed slotting and ROP-based replenishment. "
        f"The {volatile} AZ SKUs require demand buffering and escalation protocols."
    )
    ws.merge_cells("B34:I34")
    ws["B34"] = key_finding
    _style_range(ws, 34, 2, 9, fill=TITLE_FILL, font=Font(bold=True, color="FFFFFF"), alignment=LEFT)


def _write_full_sku_ranking(ws, abc_xyz: pd.DataFrame) -> None:
    widths = [3, 10, 14, 15, 10, 11, 10, 10, 10, 10, 11, 10, 12, 10, 8, 10, 22, 28]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    _apply_title(
        ws,
        "B1",
        18,
        f"FULL SKU RANKING TABLE — ABC-XYZ CLASSIFICATION ({abc_xyz['sku_code'].nunique():,} SKUs | Jul-Dec 2025)",
    )
    headers = [
        "Rank\n(Qty)",
        "SKU Code",
        "Total Qty\n(pcs)",
        "% Qty",
        "Cum %\nQty",
        "ABC\n(Qty)",
        "Rank\n(Freq)",
        "Order\nFreq",
        "% Freq",
        "Cum %\nFreq",
        "ABC\n(Freq)",
        "Total CBM",
        "CV",
        "XYZ",
        "ABC-XYZ",
        "Operational Category",
        "Slotting Zone",
    ]
    for col, value in enumerate(headers, start=2):
        ws.cell(2, col, value)
    _style_range(ws, 2, 2, 18, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)

    qty_sorted = abc_xyz.sort_values(["quantity", "sku_code"], ascending=[False, True]).reset_index(drop=True)
    qty_rank = {sku: idx for idx, sku in enumerate(qty_sorted["sku_code"], start=1)}
    freq_sorted = abc_xyz.sort_values(["order_frequency", "sku_code"], ascending=[False, True]).reset_index(drop=True)
    freq_rank = {sku: idx for idx, sku in enumerate(freq_sorted["sku_code"], start=1)}

    for row_idx, row in enumerate(qty_sorted.itertuples(index=False), start=3):
        values = [
            qty_rank[row.sku_code],
            row.sku_code,
            _safe_float(row.quantity),
            _safe_float(row.quantity_share),
            _safe_float(row.quantity_cumulative_share),
            row.abc_quantity,
            freq_rank[row.sku_code],
            _safe_float(row.order_frequency),
            _safe_float(row.frequency_share),
            _safe_float(row.frequency_cumulative_share),
            row.abc_frequency,
            _safe_float(row.cbm_total),
            _safe_float(row.demand_cv),
            row.xyz,
            row.abc_xyz,
            _operational_category(row.abc_xyz),
            _slotting_zone(row.abc_quantity),
        ]
        for col, value in enumerate(values, start=2):
            ws.cell(row_idx, col, value)
        _style_range(ws, row_idx, 2, 18, fill=CLASS_FILLS.get(row.abc_quantity), alignment=CENTER)
        ws.cell(row_idx, 3).alignment = LEFT
        ws.cell(row_idx, 17).alignment = LEFT
        ws.cell(row_idx, 18).alignment = LEFT
        ws.cell(row_idx, 5).number_format = "0.00%"
        ws.cell(row_idx, 6).number_format = "0.0%"
        ws.cell(row_idx, 10).number_format = "0.00%"
        ws.cell(row_idx, 11).number_format = "0.0%"
        ws.cell(row_idx, 14).number_format = "0.000"
    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"B2:R{ws.max_row}"


def _write_monthly_demand_sheet(ws, monthly_demand: pd.DataFrame) -> None:
    ab = monthly_demand[monthly_demand["abc_quantity"].isin(["A", "B"])].copy()
    month_columns = _month_columns(monthly_demand)
    widths = [3, 14, 8, 8] + [11] * len(month_columns) + [12, 11, 11, 8, 10]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    _apply_title(ws, "B1", 4 + len(month_columns) + 5, "MONTHLY DEMAND TABLE — A & B CLASS SKUs (XYZ Calculation Base)")
    headers = ["SKU Code", "ABC", "XYZ", *month_columns, "Total", "Mean/Mo", "Std Dev", "CV", "ABC-XYZ"]
    for col, value in enumerate(headers, start=2):
        ws.cell(2, col, value)
    _style_range(ws, 2, 2, 1 + len(headers), fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    for row_idx, (_, row) in enumerate(ab.iterrows(), start=3):
        values = [
            row["sku_code"],
            row["abc_quantity"],
            row["xyz"],
            *[row[column] for column in month_columns],
            row["Total"],
            row["Mean/Mo"],
            row["Std Dev"],
            row["CV"],
            row["abc_xyz"],
        ]
        for col, value in enumerate(values, start=2):
            ws.cell(row_idx, col, value)
        _style_range(ws, row_idx, 2, 1 + len(headers), fill=CLASS_FILLS.get(row["abc_quantity"]), alignment=CENTER)
        ws.cell(row_idx, 2).alignment = LEFT
        total_col = 2 + 3 + len(month_columns)
        mean_col = total_col + 1
        std_col = total_col + 2
        cv_col = total_col + 3
        ws.cell(row_idx, mean_col).number_format = "0.0"
        ws.cell(row_idx, std_col).number_format = "0.0"
        ws.cell(row_idx, cv_col).number_format = "0.000"
    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"B2:{get_column_letter(1 + len(headers))}{ws.max_row}"


def _write_class_a_deep_dive(ws, monthly_demand: pd.DataFrame) -> None:
    class_a = monthly_demand[monthly_demand["abc_quantity"].eq("A")].copy().sort_values("quantity", ascending=False)
    widths = [3, 8, 14, 12, 12, 12, 12, 12, 12, 12, 10, 14, 22, 20, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    _apply_title(
        ws,
        "B1",
        15,
        f"CLASS A SKUs — FAST-MOVING DEEP DIVE ({len(class_a):,} SKUs | {_format_share(class_a['quantity'].sum() / monthly_demand['quantity'].sum())} of Volume)",
    )
    ws.merge_cells("B2:O2")
    ws["B2"] = (
        "These SKUs create the highest operational pressure: prioritize pick-face slotting, cycle counting, "
        "and safety stock buffering for AZ subgroup."
    )
    ws["B2"].alignment = LEFT
    ws["B2"].font = Font(italic=True, color="5B5B5B")
    headers = [
        "Rank",
        "SKU Code",
        "ABC-XYZ\nGroup",
        "Total Qty\n(pcs)",
        "% of Total\nQty",
        "Order\nFrequency",
        "% of Total\nFreq",
        "Total CBM",
        "Avg Monthly\nDemand",
        "CV",
        "Demand\nPattern",
        "Stock\nStrategy",
        "Slotting\nPriority",
        "Reorder\nLogic",
    ]
    for col, value in enumerate(headers, start=2):
        ws.cell(3, col, value)
    _style_range(ws, 3, 2, 15, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    total_qty = monthly_demand["quantity"].sum()
    total_freq = monthly_demand["order_frequency"].sum()
    for row_idx, (_, row) in enumerate(class_a.iterrows(), start=4):
        values = [
            row_idx - 3,
            row["sku_code"],
            row["abc_xyz"],
            _safe_float(row["quantity"]),
            _safe_float(row["quantity"]) / total_qty if total_qty else 0.0,
            _safe_float(row["order_frequency"]),
            _safe_float(row["order_frequency"]) / total_freq if total_freq else 0.0,
            _safe_float(row["cbm_total"]),
            _safe_float(row["Mean/Mo"]),
            _safe_float(row["CV"]),
            _demand_pattern(row["xyz"]),
            _stock_strategy(row["xyz"]),
            "Ground Level - Zone 1",
            _reorder_logic(row["xyz"]),
        ]
        for col, value in enumerate(values, start=2):
            ws.cell(row_idx, col, value)
        _style_range(ws, row_idx, 2, 15, fill=CLASS_FILLS.get(row["xyz"]), alignment=CENTER)
        ws.cell(row_idx, 3).alignment = LEFT
        ws.cell(row_idx, 6).number_format = "0.00%"
        ws.cell(row_idx, 8).number_format = "0.00%"
        ws.cell(row_idx, 10).number_format = "0.0"
        ws.cell(row_idx, 11).number_format = "0.000"
        ws.cell(row_idx, 12).alignment = LEFT
        ws.cell(row_idx, 13).alignment = LEFT
        ws.cell(row_idx, 14).alignment = LEFT
        ws.cell(row_idx, 15).alignment = LEFT
    ws.freeze_panes = "B4"
    ws.auto_filter.ref = f"B3:O{ws.max_row}"
