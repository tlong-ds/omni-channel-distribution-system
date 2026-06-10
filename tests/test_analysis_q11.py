import sys
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

sys.path.append(str(Path(__file__).resolve().parents[1]))

from src.logage2026.analysis import (
    build_abc_xyz,
    build_abc_xyz_matrix_summary,
    build_q11_monthly_demand_table,
    filter_q11_shipments,
)
from src.logage2026.loading import load_sku_master, load_transactions
from src.logage2026.cleaning import clean_sku_master
# write_q11_workbook removed


def _minimal_q11_shipments(raw: pd.DataFrame) -> pd.DataFrame:
    frame = raw.copy()
    frame["sku_code"] = frame["SKU Code (CMMF)"].astype("string").str.strip().str.replace(r"\.0$", "", regex=True)
    frame["document_no"] = frame["Document No."].astype("string").str.strip().fillna("unknown")
    frame["source_warehouse"] = frame["Source Warehouse"].astype("string").str.strip().fillna("unknown")
    frame["order_id"] = frame["source_warehouse"] + "-" + frame["document_no"]
    frame["document_type"] = frame["Document Type"].astype("string").str.strip().str.upper()
    frame["quantity"] = pd.to_numeric(frame["Quantity"], errors="coerce")
    frame["cbm_total"] = pd.to_numeric(frame["CBM Total"], errors="coerce")
    frame["created_date"] = pd.to_datetime(frame["Created Date"], errors="coerce")
    frame["data_error_flag"] = frame["quantity"].isna() | frame["cbm_total"].isna()
    return frame[
        [
            "sku_code",
            "document_no",
            "source_warehouse",
            "order_id",
            "document_type",
            "quantity",
            "cbm_total",
            "created_date",
            "data_error_flag",
        ]
    ].dropna(subset=["sku_code", "created_date"])


def test_filter_q11_shipments_uses_configured_date_and_document_types() -> None:
    shipments = pd.DataFrame(
        [
            {
                "sku_code": "SKU-A",
                "document_type": "A/R INVOICE",
                "created_date": pd.Timestamp("2025-06-01"),
                "data_error_flag": False,
            },
            {
                "sku_code": "SKU-B",
                "document_type": "INVENTORY TRANSFERS",
                "created_date": pd.Timestamp("2025-07-15"),
                "data_error_flag": False,
            },
            {
                "sku_code": "SKU-C",
                "document_type": "SALES ORDER",
                "created_date": pd.Timestamp("2025-08-01"),
                "data_error_flag": False,
            },
            {
                "sku_code": "SKU-D",
                "document_type": "GOODS ISSUE",
                "created_date": pd.Timestamp("2026-01-01"),
                "data_error_flag": False,
            },
            {
                "sku_code": "SKU-E",
                "document_type": "STOCK COUNT",
                "created_date": pd.Timestamp("2025-09-01"),
                "data_error_flag": True,
            },
        ]
    )

    filtered = filter_q11_shipments(shipments)

    assert filtered["sku_code"].tolist() == ["SKU-B"]


def test_build_abc_xyz_supports_monthly_variability_and_missing_sku_master() -> None:
    shipments = pd.DataFrame(
        [
            {"sku_code": "SKU-STABLE", "order_id": "O1", "quantity": 10.0, "cbm_total": 1.0, "created_date": pd.Timestamp("2025-06-15")},
            {"sku_code": "SKU-STABLE", "order_id": "O2", "quantity": 10.0, "cbm_total": 1.0, "created_date": pd.Timestamp("2025-07-15")},
            {"sku_code": "SKU-STABLE", "order_id": "O3", "quantity": 10.0, "cbm_total": 1.0, "created_date": pd.Timestamp("2025-08-15")},
            {"sku_code": "SKU-STABLE", "order_id": "O4", "quantity": 10.0, "cbm_total": 1.0, "created_date": pd.Timestamp("2025-09-15")},
            {"sku_code": "SKU-STABLE", "order_id": "O5", "quantity": 10.0, "cbm_total": 1.0, "created_date": pd.Timestamp("2025-10-15")},
            {"sku_code": "SKU-STABLE", "order_id": "O6", "quantity": 10.0, "cbm_total": 1.0, "created_date": pd.Timestamp("2025-11-15")},
            {"sku_code": "SKU-STABLE", "order_id": "O7", "quantity": 10.0, "cbm_total": 1.0, "created_date": pd.Timestamp("2025-12-15")},
            {"sku_code": "SKU-VOLATILE", "order_id": "O8", "quantity": 70.0, "cbm_total": 7.0, "created_date": pd.Timestamp("2025-06-15")},
            {"sku_code": "SKU-MISSING", "order_id": "O9", "quantity": 5.0, "cbm_total": 0.5, "created_date": pd.Timestamp("2025-06-20")},
        ]
    )
    sku_master = pd.DataFrame(
        [
            {"sku_code": "SKU-STABLE", "product_name": "Stable SKU", "category": "A"},
            {"sku_code": "SKU-VOLATILE", "product_name": "Volatile SKU", "category": "B"},
        ]
    )

    abc_xyz = build_abc_xyz(shipments, sku_master)
    indexed = abc_xyz.set_index("sku_code")

    assert "SKU-MISSING" in indexed.index
    assert pd.isna(indexed.loc["SKU-MISSING", "product_name"])
    assert indexed.loc["SKU-STABLE", "xyz"] == "X"
    assert indexed.loc["SKU-VOLATILE", "xyz"] == "Z"
    assert "monthly_mean_quantity" in abc_xyz.columns
    assert "monthly_nonzero_periods" in abc_xyz.columns


@pytest.fixture(scope="module")
def live_q11_outputs() -> dict[str, pd.DataFrame]:
    sku_master = clean_sku_master(load_sku_master())
    shipments = _minimal_q11_shipments(load_transactions())
    q11_shipments = filter_q11_shipments(shipments)
    abc_xyz = build_abc_xyz(q11_shipments, sku_master)
    abc_xyz_matrix = build_abc_xyz_matrix_summary(abc_xyz)
    monthly_demand = build_q11_monthly_demand_table(abc_xyz, q11_shipments)
    return {
        "q11_shipments": q11_shipments,
        "abc_xyz": abc_xyz,
        "abc_xyz_matrix": abc_xyz_matrix,
        "monthly_demand": monthly_demand,
    }


def test_q11_live_data_matches_reference_counts(live_q11_outputs: dict[str, pd.DataFrame]) -> None:
    q11_shipments = live_q11_outputs["q11_shipments"]
    abc_xyz = live_q11_outputs["abc_xyz"]
    abc_xyz_matrix = live_q11_outputs["abc_xyz_matrix"]

    assert len(q11_shipments) == 45_172
    assert abc_xyz["sku_code"].nunique() == 612
    assert abc_xyz["abc_quantity"].value_counts().sort_index().to_dict() == {"A": 41, "B": 74, "C": 497}
    assert abc_xyz["xyz"].value_counts().sort_index().to_dict() == {"X": 52, "Y": 180, "Z": 380}
    observed_matrix = {
        (row["abc_quantity"], row["xyz"]): int(row["sku_count"])
        for _, row in abc_xyz_matrix.iterrows()
    }
    assert observed_matrix == {
        ("A", "X"): 10,
        ("A", "Y"): 23,
        ("A", "Z"): 8,
        ("B", "X"): 16,
        ("B", "Y"): 34,
        ("B", "Z"): 24,
        ("C", "X"): 26,
        ("C", "Y"): 123,
        ("C", "Z"): 348,
    }


def test_q11_workbook_smoke() -> None:
    output_path = Path("outputs/round2/summary_tables.xlsx")
    if not output_path.exists():
        from run_analysis import main
        main()

    workbook = load_workbook(output_path, data_only=True)
    expected_sheets = [
        "Q1.1 Dashboard",
        "Q1.1 Full SKU Ranking",
        "Q1.1 Monthly Demand",
        "Q1.1 Class A Deep Dive",
        "Q1.2 Warehouse by Region",
        "Q1.2 Top Provinces",
        "Q1.2 Warehouse Imbalance",
        "Q1.3 Segment Profile",
        "Q1.3 Packaging",
        "Q1.3 Geo Spread",
        "Q2.2 Safety Stock",
        "Q2.2 Lead Time Sensitivity",
        "Q2.2 Inventory Pooling",
        "Q2.1 HCM Districts",
        "Q2.1 Dark Store SLA",
    ]
    for sheet in expected_sheets:
        assert sheet in workbook.sheetnames

    dashboard = workbook["Q1.1 Dashboard"]
    ranking = workbook["Q1.1 Full SKU Ranking"]
    monthly = workbook["Q1.1 Monthly Demand"]
    class_a = workbook["Q1.1 Class A Deep Dive"]

    assert dashboard["B1"].value == "LOGage 2026 — QUESTION 1.1 | ABC-XYZ ANALYSIS DASHBOARD"
    assert dashboard["C6"].value == 41
    assert dashboard["C12"].value == 52
    assert ranking.max_row == 614
    assert ranking["C3"].value == "4200009163"
    assert ranking["D3"].value == 39523
    assert ranking["O3"].value == "X"
    assert monthly.max_row == 117
    assert monthly["B3"].value == "4200009163"
    assert monthly["E3"].value == 5588
    assert class_a["C4"].value == "4200009163"
    assert class_a["D4"].value == "AX"
