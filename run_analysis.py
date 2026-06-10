from src.logage2026.analysis import (
    build_abc_xyz,
    build_abc_xyz_matrix_summary,
    build_classification_metadata,
    build_customer_cluster_summary,
    build_customer_match_quality_summary,
    build_document_type_summary,
    build_fast_moving_summary,
    build_geography_coverage_summary,
    build_geography_diagnostics_summary,
    build_geography_source_summary,
    build_missing_data_summary,
    build_hcm_district_summary,
    build_inventory_pooling_summary,
    build_lead_time_region_table,
    build_lead_time_sensitivity,
    build_network_model_evaluation,
    build_q21_channel_flow_summary,
    build_order_profile_segments,
    build_q11_monthly_demand_table,
    build_q12_province_correlation_input_summary,
    build_q12_province_demand_summary,
    build_q12_province_cluster_summary,
    build_q12_province_warehouse_dominance_summary,
    build_q12_region_orders_quantity_summary,
    build_q12_urban_provincial_summary,
    build_q12_warehouse_imbalance_visual_summary,
    build_q13_segment_geographic_spread_summary,
    build_q13_segment_packaging_summary,
    build_q13_segment_profile_summary,
    build_q13_segment_province_spread_summary,
    build_safety_stock_class_a,
    build_sku_pick_profile,
    build_slotting_plan,
    build_unresolved_candidate_region_summary,
    build_unresolved_customer_summary,
    build_warehouse_imbalance_summary,
    build_warehouse_region_summary,
    compute_travel_time_metrics,
    filter_assignment_shipments,
    filter_q11_shipments,
)
from src.logage2026.cleaning import clean_distributors, clean_shipments, clean_sku_master
from src.logage2026.config import CHARTS_DIR, CLEANED_DIR, NOTES_DIR, OUTPUT_DIR, Q11_WORKBOOK_OUTPUT, TABLES_DIR
from src.logage2026.loading import (
    load_distributors,
    load_segment_overrides,
    load_sku_master,
    load_transactions,
)
from src.logage2026.notes import write_notes, write_part3_notes
from src.logage2026.excel_reports import write_summary_workbook
from src.logage2026.visuals import boundary_province_names, save_charts, save_q31_u_shape_heatmap, save_q32_flowchart_image


EXPECTED_ASSIGNMENT_ROWS = 43_894
EXPECTED_ASSIGNMENT_QUANTITY = 355_364.80
EXPECTED_ASSIGNMENT_CBM = 19_653.78
EXPECTED_Q11_ROWS = 45_172
EXPECTED_Q11_SKUS = 612
EXPECTED_Q11_ABC_COUNTS = {"A": 41, "B": 74, "C": 497}
EXPECTED_Q11_XYZ_COUNTS = {"X": 52, "Y": 180, "Z": 380}
NOTE_FILENAME = "part1_question_summary.tex"

STALE_OUTPUTS = [
    TABLES_DIR / "safety_stock_class_a.csv",
    NOTES_DIR / "recommendations.md",
    NOTES_DIR / "question_summary.md",
    TABLES_DIR / "abc_xyz.csv",
    TABLES_DIR / "abc_xyz_matrix_summary.csv",
    TABLES_DIR / "fast_moving_summary.csv",
    TABLES_DIR / "classification_metadata.csv",
    TABLES_DIR / "missing_data_summary.csv",
    TABLES_DIR / "geography_coverage_summary.csv",
    TABLES_DIR / "geography_diagnostics_summary.csv",
    TABLES_DIR / "unresolved_candidate_region_summary.csv",
    TABLES_DIR / "warehouse_region_summary.csv",
    TABLES_DIR / "customer_cluster_summary.csv",
    TABLES_DIR / "warehouse_imbalance_summary.csv",
    TABLES_DIR / "q12_region_orders_quantity_summary.csv",
    TABLES_DIR / "q12_province_cluster_summary.csv",
    TABLES_DIR / "order_profile_segments.csv",
    TABLES_DIR / "document_type_summary.csv",
    TABLES_DIR / "customer_match_quality_summary.csv",
    TABLES_DIR / "geography_source_summary.csv",
    TABLES_DIR / "unresolved_customer_summary.csv",
    CHARTS_DIR / "abc_quantity_distribution.png",
    CHARTS_DIR / "abc_xyz_matrix.png",
    CHARTS_DIR / "regional_quantity_density.png",
    CHARTS_DIR / "warehouse_region_split.png",
    CHARTS_DIR / "q12_province_demand_maps.png",
    CHARTS_DIR / "order_profile_comparison.png",
    CHARTS_DIR / "vietnam_regions_map.png",
]


def main() -> None:
    for directory in [CLEANED_DIR, TABLES_DIR, NOTES_DIR]:
        directory.mkdir(parents=True, exist_ok=True)

    _remove_stale_outputs()

    sku_master = clean_sku_master(load_sku_master())
    distributors = clean_distributors(load_distributors())
    segment_overrides = load_segment_overrides()
    shipments = clean_shipments(load_transactions(), distributors, segment_overrides=segment_overrides)
    q11_shipments = filter_q11_shipments(shipments)
    assignment_shipments = filter_assignment_shipments(shipments)

    abc_xyz = build_abc_xyz(q11_shipments, sku_master)
    abc_xyz_matrix_frequency = build_abc_xyz_matrix_summary(abc_xyz, "abc_frequency")
    abc_xyz_matrix_volatility = build_abc_xyz_matrix_summary(abc_xyz, "xyz_volatility")

    fast_moving_summary = build_fast_moving_summary(abc_xyz)
    classification_metadata = build_classification_metadata()
    q11_monthly_demand = build_q11_monthly_demand_table(abc_xyz, q11_shipments)
    missing_data_summary = build_missing_data_summary(shipments)
    geography_coverage_summary = build_geography_coverage_summary(assignment_shipments)
    geography_diagnostics_summary = build_geography_diagnostics_summary(assignment_shipments)
    unresolved_candidate_region_summary = build_unresolved_candidate_region_summary(assignment_shipments)
    warehouse_region_summary = build_warehouse_region_summary(assignment_shipments)
    customer_cluster_summary = build_customer_cluster_summary(assignment_shipments)
    warehouse_imbalance_summary = build_warehouse_imbalance_summary(assignment_shipments)
    q12_region_orders_quantity_summary = build_q12_region_orders_quantity_summary(assignment_shipments)
    q12_province_cluster_summary = build_q12_province_cluster_summary(assignment_shipments)
    q12_province_demand_summary = build_q12_province_demand_summary(assignment_shipments)
    q12_province_warehouse_dominance_summary = build_q12_province_warehouse_dominance_summary(assignment_shipments)
    q12_province_correlation_input_summary = build_q12_province_correlation_input_summary(assignment_shipments)
    q12_urban_provincial_summary = build_q12_urban_provincial_summary(assignment_shipments)
    q12_warehouse_imbalance_visual_summary = build_q12_warehouse_imbalance_visual_summary(assignment_shipments)
    q13_segment_profile_summary = build_q13_segment_profile_summary(assignment_shipments, sku_master)
    q13_segment_packaging_summary = build_q13_segment_packaging_summary(assignment_shipments, sku_master)
    q13_segment_geographic_spread_summary = build_q13_segment_geographic_spread_summary(assignment_shipments)
    q13_segment_province_spread_summary = build_q13_segment_province_spread_summary(assignment_shipments)
    order_profile_segments = build_order_profile_segments(assignment_shipments, sku_master)
    document_type_summary = build_document_type_summary(shipments)
    customer_match_quality_summary = build_customer_match_quality_summary(assignment_shipments)
    geography_source_summary = build_geography_source_summary(assignment_shipments)
    unresolved_customer_summary = build_unresolved_customer_summary(assignment_shipments)

    safety_stock_class_a = build_safety_stock_class_a(shipments, abc_xyz)
    lead_time_sensitivity = build_lead_time_sensitivity(safety_stock_class_a)
    inventory_pooling_summary = build_inventory_pooling_summary(safety_stock_class_a)
    lead_time_table = build_lead_time_region_table()
    hcm_district_summary = build_hcm_district_summary(shipments)
    network_model_evaluation = build_network_model_evaluation(hcm_district_summary)
    q21_channel_flow_summary = build_q21_channel_flow_summary(shipments)

    # Part 3 metrics needed early for summary workbook
    sku_pick_profile = build_sku_pick_profile(shipments, sku_master, abc_xyz)
    slotting_plan = build_slotting_plan(shipments, sku_master, abc_xyz)
    travel_metrics = compute_travel_time_metrics(abc_xyz)

    sku_master.to_csv(CLEANED_DIR / "sku_master_cleaned.csv", index=False)
    distributors.to_csv(CLEANED_DIR / "distributors_cleaned.csv", index=False)
    shipments.to_csv(CLEANED_DIR / "shipments_cleaned.csv", index=False)

    abc_xyz.to_csv(TABLES_DIR / "q11_sku_abc_xyz.csv", index=False)
    abc_xyz_matrix_frequency.to_csv(TABLES_DIR / "q11_abc_xyz_matrix_frequency_summary.csv", index=False)
    abc_xyz_matrix_volatility.to_csv(TABLES_DIR / "q11_abc_xyz_matrix_volatility_summary.csv", index=False)

    fast_moving_summary.to_csv(TABLES_DIR / "q11_fast_moving_summary.csv", index=False)
    classification_metadata.to_csv(TABLES_DIR / "q11_classification_metadata.csv", index=False)
    q11_monthly_demand.to_csv(TABLES_DIR / "q11_monthly_demand_summary.csv", index=False)
    missing_data_summary.to_csv(TABLES_DIR / "shared_missing_data_summary.csv", index=False)
    geography_coverage_summary.to_csv(TABLES_DIR / "q12_geography_coverage_summary.csv", index=False)
    geography_diagnostics_summary.to_csv(TABLES_DIR / "q12_geography_diagnostics_summary.csv", index=False)
    unresolved_candidate_region_summary.to_csv(TABLES_DIR / "q12_unresolved_candidate_region_summary.csv", index=False)
    warehouse_region_summary.to_csv(TABLES_DIR / "q12_warehouse_region_summary.csv", index=False)
    customer_cluster_summary.to_csv(TABLES_DIR / "q12_customer_cluster_summary.csv", index=False)
    warehouse_imbalance_summary.to_csv(TABLES_DIR / "q12_warehouse_imbalance_summary.csv", index=False)
    q12_region_orders_quantity_summary.to_csv(TABLES_DIR / "q12_region_quantity_orders_summary.csv", index=False)
    q12_province_cluster_summary.to_csv(TABLES_DIR / "q12_top_demand_provinces_summary.csv", index=False)
    q12_province_demand_summary.to_csv(TABLES_DIR / "q12_province_demand_summary.csv", index=False)
    q12_province_warehouse_dominance_summary.to_csv(
        TABLES_DIR / "q12_province_warehouse_dominance_summary.csv", index=False
    )
    q12_province_correlation_input_summary.to_csv(
        TABLES_DIR / "q12_province_correlation_input_summary.csv", index=False
    )
    q12_urban_provincial_summary.to_csv(TABLES_DIR / "q12_urban_provincial_summary.csv", index=False)
    q12_warehouse_imbalance_visual_summary.to_csv(TABLES_DIR / "q12_warehouse_imbalance_visual_summary.csv", index=False)
    q13_segment_profile_summary.to_csv(TABLES_DIR / "q13_segment_profile_summary.csv", index=False)
    q13_segment_packaging_summary.to_csv(TABLES_DIR / "q13_segment_packaging_summary.csv", index=False)
    q13_segment_geographic_spread_summary.to_csv(TABLES_DIR / "q13_segment_geographic_spread_summary.csv", index=False)
    q13_segment_province_spread_summary.to_csv(TABLES_DIR / "q13_segment_province_spread_summary.csv", index=False)
    order_profile_segments.to_csv(TABLES_DIR / "q13_order_profile_segments.csv", index=False)
    document_type_summary.to_csv(TABLES_DIR / "shared_document_type_summary.csv", index=False)
    customer_match_quality_summary.to_csv(TABLES_DIR / "q12_customer_match_quality_summary.csv", index=False)
    geography_source_summary.to_csv(TABLES_DIR / "q12_geography_source_summary.csv", index=False)
    unresolved_customer_summary.to_csv(TABLES_DIR / "q12_unresolved_customer_summary.csv", index=False)
    
    safety_stock_class_a.to_csv(TABLES_DIR / "safety_stock_class_a.csv", index=False)
    lead_time_sensitivity.to_csv(TABLES_DIR / "lead_time_sensitivity.csv", index=False)
    inventory_pooling_summary.to_csv(TABLES_DIR / "inventory_pooling_summary.csv", index=False)
    hcm_district_summary.to_csv(TABLES_DIR / "hcm_district_summary.csv", index=False)
    network_model_evaluation.to_csv(TABLES_DIR / "q21_network_model_evaluation.csv", index=False)
    q21_channel_flow_summary.to_csv(TABLES_DIR / "q21_channel_flow_summary.csv", index=False)
    
    import pandas as pd
    import openpyxl
    from openpyxl.utils import get_column_letter
    
    print("Exporting cleaned data to cleaned_data.xlsx...")
    cleaned_path = CLEANED_DIR / "cleaned_data.xlsx"
    with pd.ExcelWriter(cleaned_path, engine='openpyxl') as writer:
        pipeline_summary_data = [
            {"Dataset": "SKU Master", "Pipeline Stage": "Standardization", "Details": "Stripped whitespace, converted missing values to proper nulls, normalized strings to uppercase."},
            {"Dataset": "SKU Master", "Pipeline Stage": "Volume Derivation", "Details": "Calculated carton volume in CBM strictly from L x W x H dimensions if missing."},
            {"Dataset": "SKU Master", "Pipeline Stage": "Pallet Capacity & Loose Items", "Details": "Calculated loose pieces per pallet via residual of total pieces minus full carton capacity."},
            {"Dataset": "SKU Master", "Pipeline Stage": "Weight Imputation", "Details": "Derived missing carton weights from piece weights (and vice versa) using pieces per carton. Flagged conflicts."},
            {"Dataset": "SKU Master", "Pipeline Stage": "Deduplication", "Details": "Kept first unique row per SKU Code to eliminate redundant master records."},
            {"Dataset": "Distributor Network", "Pipeline Stage": "Normalization", "Details": "Normalized customer names, stripping legal entity prefixes for stable key matching."},
            {"Dataset": "Distributor Network", "Pipeline Stage": "Address Parsing", "Details": "Parsed raw delivery addresses using regex and dictionaries to extract Base Address, Province, District, and Ward."},
            {"Dataset": "Distributor Network", "Pipeline Stage": "Deduplication", "Details": "Removed ambiguous overlapping locations based on exact lat/long and text similarity matching."},
            {"Dataset": "Distributor Network", "Pipeline Stage": "Distance Calculation", "Details": "Computed straight-line geographical distances from main hubs (My Phuoc, Vinh Loc) using Haversine formula."},
            {"Dataset": "Shipment Transactions", "Pipeline Stage": "Order Context", "Details": "Concatenated Source Warehouse and Document No. to uniquely identify order batches."},
            {"Dataset": "Shipment Transactions", "Pipeline Stage": "Channel Inference", "Details": "Classified channel as B2B or B2C based on textual regex patterns in the Ship-To customer name."},
            {"Dataset": "Shipment Transactions", "Pipeline Stage": "Distributor Matching", "Details": "Joined shipments with cleaned Distributor lookup on normalized customer keys to inherit reliable geocoding."},
            {"Dataset": "Shipment Transactions", "Pipeline Stage": "Address Fallback", "Details": "For unmatched customers, parsed the transaction 'Ship-to' text directly to extract geographical components."},
            {"Dataset": "Shipment Transactions", "Pipeline Stage": "Segment Mapping", "Details": "Categorized customers into segments (Modern vs Traditional Trade) using keyword rules and external overrides."}
        ]
        summary_df = pd.DataFrame(pipeline_summary_data)
        summary_df.to_excel(writer, sheet_name="Pipeline_Transparency", index=False)
        
        for csv_file in CLEANED_DIR.glob("*.csv"):
            df = pd.read_csv(csv_file)
            df.to_excel(writer, sheet_name=csv_file.stem[:31], index=False)
            
    # Apply formatting
    wb = openpyxl.load_workbook(cleaned_path)
    
    # Format Pipeline_Transparency
    if 'Pipeline_Transparency' in wb.sheetnames:
        ws = wb['Pipeline_Transparency']
        from openpyxl.styles import Alignment
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 80
        
        align_center = Alignment(wrap_text=True, vertical='center', horizontal='center')
        align_top_left = Alignment(wrap_text=True, vertical='center', horizontal='left')
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            if row[0].row == 1:
                for cell in row:
                    cell.alignment = align_center
                continue
            
            row[0].alignment = align_center
            row[1].alignment = align_top_left
            if len(row) > 2:
                row[2].alignment = align_top_left
                details_text = str(row[2].value) if row[2].value else ""
                num_lines = (len(details_text) // 70) + 1
                ws.row_dimensions[row[0].row].height = num_lines * 18
                
        start_row = 2
        current_val = ws.cell(row=start_row, column=1).value
        for r in range(3, ws.max_row + 2):
            cell_val = ws.cell(row=r, column=1).value if r <= ws.max_row else None
            if cell_val != current_val:
                if r - 1 > start_row:
                    ws.merge_cells(start_row=start_row, start_column=1, end_row=r-1, end_column=1)
                start_row = r
                current_val = cell_val
    
    # Format distributors
    if 'distributors_cleaned' in wb.sheetnames:
        ws = wb['distributors_cleaned']
        keep_letters = {'A', 'D', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O'}
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            if letter not in keep_letters:
                ws.column_dimensions[letter].hidden = True
                
    # Format sku_master
    if 'sku_master_cleaned' in wb.sheetnames:
        ws = wb['sku_master_cleaned']
        keep_letters = {'A', 'E', 'F', 'H'}
        for c in range(ord('I'), ord('Z')+1):
            keep_letters.add(chr(c))
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            if letter not in keep_letters:
                ws.column_dimensions[letter].hidden = True
                
    # Format shipments
    if 'shipments_cleaned' in wb.sheetnames:
        ws = wb['shipments_cleaned']
        keep_letters = {'B', 'C', 'F', 'G', 'H', 'I', 'J', 'N', 'U', 'X', 'Y', 'Z', 'AA', 'AB', 'AE', 'AF', 'AG', 'AH', 'AI'}
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            if letter not in keep_letters:
                ws.column_dimensions[letter].hidden = True

    # Auto-fit all sheets
    for ws in wb.worksheets:
        if ws.title == 'Pipeline_Transparency':
            continue
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            if ws.column_dimensions[letter].hidden:
                continue
            max_length = 0
            for cell in ws[letter][0:1000]: 
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(max_length + 2, 50)
            
    wb.save(cleaned_path)
            
    print("Exporting summary tables to summary_table.xlsx...")
    write_summary_workbook(
        abc_xyz=abc_xyz,
        abc_xyz_matrix=abc_xyz_matrix_frequency,
        monthly_demand=q11_monthly_demand,
        q11_shipments=q11_shipments,
        warehouse_region_summary=warehouse_region_summary,
        q12_top_demand_provinces_summary=q12_province_cluster_summary,
        warehouse_imbalance_summary=warehouse_imbalance_summary,
        q13_segment_profile_summary=q13_segment_profile_summary,
        q13_segment_packaging_summary=q13_segment_packaging_summary,
        q13_segment_geographic_spread_summary=q13_segment_geographic_spread_summary,
        safety_stock_class_a=safety_stock_class_a,
        lead_time_sensitivity=lead_time_sensitivity,
        inventory_pooling_summary=inventory_pooling_summary,
        hcm_district_summary=hcm_district_summary,
        network_model_evaluation=network_model_evaluation,
        shipments=shipments,
        slotting_plan=slotting_plan,
        sku_pick_profile=sku_pick_profile,
        travel_metrics=travel_metrics,
        lead_time_table=lead_time_table,
        output_path=OUTPUT_DIR / "summary_table.xlsx",
    )

    save_charts(
        assignment_shipments,
        abc_xyz,
        abc_xyz_matrix_frequency,
        abc_xyz_matrix_volatility,
        warehouse_region_summary,
        q12_region_orders_quantity_summary,
        q12_province_cluster_summary,
        q12_province_demand_summary,
        q12_province_warehouse_dominance_summary,
        q12_province_correlation_input_summary,
        q12_urban_provincial_summary,
        q12_warehouse_imbalance_visual_summary,
        q13_segment_profile_summary,
        q13_segment_packaging_summary,
        q13_segment_geographic_spread_summary,
        q13_segment_province_spread_summary,
        sku_master,
        safety_stock_class_a,
        lead_time_sensitivity,
        inventory_pooling_summary,
        hcm_district_summary,
        network_model_evaluation,
        q21_channel_flow_summary,
    )
    write_notes(
        assignment_shipments,
        abc_xyz,
        abc_xyz_matrix_frequency,
        abc_xyz_matrix_volatility,
        fast_moving_summary,
        classification_metadata,
        missing_data_summary,
        geography_coverage_summary,
        warehouse_region_summary,
        customer_cluster_summary,
        warehouse_imbalance_summary,
        q12_region_orders_quantity_summary,
        q12_province_cluster_summary,
        q12_province_demand_summary,
        q12_province_warehouse_dominance_summary,
        q12_urban_provincial_summary,
        q12_warehouse_imbalance_visual_summary,
        q13_segment_profile_summary,
        q13_segment_packaging_summary,
        q13_segment_geographic_spread_summary,
        safety_stock_class_a,
        lead_time_sensitivity,
        inventory_pooling_summary,
        hcm_district_summary,
        network_model_evaluation,
        q21_channel_flow_summary,
    )
    print(f"Round 2 analysis written to {OUTPUT_DIR}")
    from src.logage2026.analysis import Q11_END, Q11_START, ASSIGNMENT_START, ASSIGNMENT_END
    print(
        "Q1.1 workbook window: "
        f"{Q11_START.date().isoformat()} to {Q11_END.date().isoformat()} "
        f"| rows: {len(q11_shipments):,} "
        f"| classified SKUs: {abc_xyz['sku_code'].nunique():,}"
    )
    print(
        "Assignment window: "
        f"{ASSIGNMENT_START.date().isoformat()} to {ASSIGNMENT_END.date().isoformat()} "
        f"| rows: {len(assignment_shipments):,} "
        f"| quantity: {assignment_shipments['quantity'].sum():,.2f} "
        f"| CBM: {assignment_shipments['cbm_total'].sum():,.2f}"
    )
    print(f"Classified SKUs: {abc_xyz['sku_code'].nunique():,}")

    # ── Part 3 ────────────────────────────────────────────────────────────────
    print("Building Part 3: Slotting Optimization ...")
    sku_pick_profile.to_csv(TABLES_DIR / "q31_sku_pick_profile.csv", index=False)
    slotting_plan.to_csv(TABLES_DIR / "slotting_plan.csv", index=False)

    print("Rendering Part 3 charts ...")
    save_q31_u_shape_heatmap()
    save_q32_flowchart_image()

    print("Writing Part 3 LaTeX report ...")
    write_part3_notes(shipments, sku_master, abc_xyz)
    
    print("Cleaning up temporary LaTeX files (.aux, .log, etc.) in notes/ ...")
    for pattern in ["*.aux", "*.log", "*.toc", "*.out"]:
        for temp_file in NOTES_DIR.glob(pattern):
            try:
                temp_file.unlink()
            except Exception as e:
                print(f"Warning: could not delete {temp_file}: {e}")
                
    print("Part 3 complete.")


def _remove_stale_outputs() -> None:
    for path in STALE_OUTPUTS:
        if path.exists():
            path.unlink()


if __name__ == "__main__":
    main()
